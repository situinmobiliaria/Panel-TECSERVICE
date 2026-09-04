#!/usr/bin/env python3
"""
extractor.py — Genera dashboard_contratos_v16.2.html directamente desde Excel.

Uso:
    python extractor.py

Requiere:
    pip install openpyxl pandas

Configuración:
    Editar la sección CONFIG si cambia el presupuesto anual o el nombre del Excel.
"""
from __future__ import annotations
import os, re, json, math, unicodedata
from datetime import date, datetime, timedelta
from collections import defaultdict

import openpyxl
import pandas as pd

def _xlsb_to_xlsx(xlsb_path):
    """Convierte .xlsb a .xlsx temporal usando Excel COM. Retorna la ruta del xlsx generado."""
    import win32com.client, shutil, tempfile
    # Nombre único por corrida: evita PermissionError si una corrida previa
    # dejó el archivo bloqueado por un proceso Excel colgado.
    tmp_dir  = tempfile.gettempdir()
    tmp_path = os.path.join(tmp_dir, f"panel_ts_conv_{os.getpid()}.xlsx")
    # Limpiar conversiones viejas que ya no estén bloqueadas
    for f in os.listdir(tmp_dir):
        if f.startswith("panel_ts_conv") and f.endswith(".xlsx"):
            try: os.remove(os.path.join(tmp_dir, f))
            except OSError: pass
    print(f"  Convirtiendo {os.path.basename(xlsb_path)} con Excel COM ...")
    excel = None
    wb_com = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        try: excel.Visible = False
        except Exception: pass
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        wb_com = excel.Workbooks.Open(
            os.path.abspath(xlsb_path),
            UpdateLinks=0, ReadOnly=True
        )
        wb_com.SaveAs(tmp_path, FileFormat=51)  # 51 = xlOpenXMLWorkbook
        print(f"  Conversion completada -> {tmp_path}")
        return tmp_path
    finally:
        if wb_com:
            try: wb_com.Close(False)
            except: pass
        if excel:
            try: excel.Quit()
            except: pass

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG — editar aquí si cambia algo
# ══════════════════════════════════════════════════════════════════════════════
DIR  = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(DIR)
_BASE_NAME = "CONTRATOS -FACTURACION -SATISFACCION -VISITAS"
XLSX  = os.path.join(ROOT, "data", _BASE_NAME + ".xlsx")
XLSB  = os.path.join(ROOT, "data", _BASE_NAME + ".xlsb")
TMPL  = os.path.join(DIR, "template.html")
OUT   = os.path.join(DIR, "dashboard_contratos_v16.2.html")

PPTO_ANUAL_TOTAL = 2_724_000_000   # Presupuesto anual total del área (CLP)

EJECUTIVOS_VISITAS = ["Eglys Ramirez", "Cristian Perez"]

# IMPORTANTE: esta lista debe estar sincronizada con $jsFiles en build.ps1.
# Si un archivo falta aquí, esa hoja simplemente no existe en el standalone
# (dashboard_contratos_v16.2.html) aunque sí funcione en index.html. Faltaban
# base_instalada, mapa, matriz, casos y pdf, así que esas 5 hojas nunca se
# generaron en el standalone. main() ahora avisa si las listas divergen.
JS_FILES = [
    "utils.js", "datos.js", "hoja_resumen.js", "hoja_tipos.js", "hoja_nuevos.js",
    "hoja_vencimientos.js", "hoja_vision.js", "hoja_presupuesto.js",
    "hoja_facturacion.js", "hoja_panelfact.js", "hoja_base_instalada.js", "hoja_satisfaccion.js",
    "hoja_visitas.js", "hoja_mapa.js", "hoja_matriz.js", "hoja_casos.js", "hoja_alerta.js",
    "hoja_pdf.js", "hoja_eerr.js", "hoja_desglose.js", "hoja_inv_ts.js", "hoja_rep_vend.js",
    "hoja_prosp_bi.js", "hoja_cli_rel.js", "hoja_pipeline.js", "hoja_brechas.js",
]

ANO   = date.today().year
TODAY = date.today()

# ══════════════════════════════════════════════════════════════════════════════
# Clasificación de contratos por línea de negocio (Esterilización = base/default).
# Fuente Dental: Excel "CONTRATOS DENTAL" (col Línea de Negocio = DENTAL) al 2026-07-17.
# Fuente Endoscopía: listado "Contratos Vigentes ENDO" al 2026-07-15.
# Si se agregan nuevos contratos Dental/Endoscopía, sumar aquí su número de contrato.
# ══════════════════════════════════════════════════════════════════════════════
DENTAL_CONTRATOS_NUMS = {
    163, 194, 192, 193, 213, 221, 222, 139, 160, 165, 164, 206,
    188, 232, 251, 170, 181, 183, 199, 238, 146, 187,
}
ENDOSCOPIA_CONTRATOS_NUMS = {200, 198, 142, 237}

# NOTA (2026-07-19): #198 (HUAP) y #200 (Intermedical) tienen en CONTRATOS
# TODOS fechas/Estado desactualizados (figuran Expirado, con fin 30/04/2026 y
# 31/05/2026). Cristián confirmó explícitamente que las fechas correctas son
# las del listado "Contratos Vigentes ENDO" (imagen, 2026-07-15), no las del
# Excel. Se sobreescriben inicio/fin/estado con esos valores hasta que
# CONTRATOS TODOS se actualice en la fuente (ver SUPUESTOS.txt, punto 2).
CONTRATO_OVERRIDE = {
    198: {"inicio": date(2024, 10, 31), "fin": date(2026, 10, 31), "estado": "Activado"},  # HUAP Endoscopía
    200: {"inicio": date(2026, 7, 8),   "fin": date(2027, 7, 31),  "estado": "Activado"},  # Intermedical Endoscopía
}

# Marcas con facturación propia relevante dentro del catálogo "Servicio Técnico"
# no ligado a contrato; el resto se agrupa en "Otras Marcas".
TOP_MARCAS_DESGLOSE = ["TECSERVICE", "NACIONAL", "ICTGROUP", "STEELCO", "PENTAX MEDICAL"]

def linea_negocio_contrato(num_str):
    try:
        n = int(num_str)
    except (TypeError, ValueError):
        return "Esterilización"
    if n in DENTAL_CONTRATOS_NUMS:
        return "Dental"
    if n in ENDOSCOPIA_CONTRATOS_NUMS:
        return "Endoscopía"
    return "Esterilización"

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def to_float(val, default=0.0):
    try:
        if val is None:
            return default
        if isinstance(val, float) and math.isnan(val):
            return default
        return float(val)
    except Exception:
        return default

def to_int(val, default=0):
    try:
        return int(val) if val is not None else default
    except Exception:
        return default

def safe_str(val):
    return str(val).strip() if val is not None else ""

def _bi_total(val):
    """Convierte el valor de BI Total; retorna None si es N/A o vacío."""
    if val is None:
        return None
    if isinstance(val, str) and val.strip().upper() in ('#N/A', '', 'N/A', '#REF!', '#VALUE!', '#NUM!'):
        return None
    try:
        result = int(float(val))
        return result if result > 0 else None
    except Exception:
        return None

def parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(val[:len(fmt)], fmt).date()
            except Exception:
                pass
    return None

_MESES_LARGO = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
                 "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

def formato_actualizacion(dt):
    """'Última actualización: 19 de julio 2026 · 02:50 am' a partir de un datetime."""
    mes = _MESES_LARGO[dt.month - 1]
    hora12 = dt.hour % 12 or 12
    ampm = "am" if dt.hour < 12 else "pm"
    return f"Última actualización: {dt.day} de {mes} {dt.year} · {hora12:02d}:{dt.minute:02d} {ampm}"

# ══════════════════════════════════════════════════════════════════════════════
# 1. HOJA: CONTRATOS TODOS
#    Fila 1 = números de mes (cabecera auxiliar)
#    Fila 2 = encabezados de columna
#    Fila 3+ = datos
#
#    Columnas relevantes (índice 0-based):
#    0=Coordinador  1=Vendedor  3=N°Contrato  4=Cliente  5=Garantia
#    6=FechaInicio  8=FechaFin  10=Estado
#    13-24 = flags Fact Ene–Dic
#    27=ValMes(CLP)  29=ValAnual2026  32=RealYTD2026  33=Real2025  34=Real2024
# ══════════════════════════════════════════════════════════════════════════════
def read_contratos(wb):
    ws = wb["CONTRATOS TODOS"]
    contratos = []
    last_coord   = ""
    last_vendedor = ""

    for row in ws.iter_rows(min_row=3, values_only=True):
        num_str = safe_str(row[3])
        if not num_str:
            continue

        # Forward-fill merged coordinator / vendedor cells
        coord_raw    = safe_str(row[0])
        vendedor_raw = safe_str(row[1])
        if coord_raw:
            last_coord = coord_raw
        if vendedor_raw:
            last_vendedor = vendedor_raw

        cliente = safe_str(row[4])
        if not cliente:
            continue

        es_garantia  = safe_str(row[5]).upper() in ("TRUE", "VERDADERO", "1")
        fecha_inicio = parse_date(row[6])
        fecha_fin    = parse_date(row[8])
        estado       = safe_str(row[10])   # "Activado" / "Expirado"
        try:
            override = CONTRATO_OVERRIDE.get(int(num_str))
        except ValueError:
            override = None
        if override:
            fecha_inicio = override["inicio"]
            fecha_fin    = override["fin"]
            estado       = override["estado"]

        if not fecha_inicio or not fecha_fin:
            continue

        # Billing flags per month (columns N–Y = indices 13–24)
        fact_flags = [bool(row[13 + m]) for m in range(12)]

        cuota_uf  = to_float(row[25])   # Z:  Facturación Cuota UF
        val_mes   = to_float(row[27])   # AB: Facturación Neta Mes convertido
        n_mant    = to_int(row[28])     # AC: N° mantenimientos esperados año
        val_anual = to_float(row[29])   # AE: Facturación Anual Esperada 2026
        if val_anual == 0:
            val_anual = val_mes * n_mant if n_mant >= 1 else val_mes  # fallback si col AE vacía
        n_mant_actual = to_int(row[31])   # AF: N° Mantenimientos a la Fecha 2026
        real_ytd  = to_float(row[32])   # AG: Facturación Contratos 2026 YTD
        real_2025 = to_float(row[33])   # AH
        real_2024 = to_float(row[34])   # AI
        # AJ (idx 35): tipo de programa (Basic/Advanced/Profesional/Integral Care Program)
        prog_raw  = row[35] if len(row) > 35 else None
        programa  = safe_str(prog_raw).replace('​','').strip() if prog_raw else ""
        # AK (idx 36): flag "FACTURACIÓN < CONTRATOS" (SI/NO)
        bajo_contrato = safe_str(row[36]).strip().upper() if len(row) > 36 and row[36] else ""

        dias_vence    = (fecha_fin - TODAY).days
        long_dias     = max(1, (fecha_fin - fecha_inicio).days)
        tpo_activo    = max(0, (TODAY - fecha_inicio).days)
        pct_consumido = round(tpo_activo / long_dias * 100, 1)

        contratos.append({
            "n":            num_str,
            "cliente":      cliente,
            "coord":        last_coord,
            "vendedor":     last_vendedor,
            "tipo":         "Garantia" if es_garantia else "Comercial",
            "inicio":       fecha_inicio.isoformat(),
            "fin":          fecha_fin.isoformat(),
            "inicio_fmt":   fecha_inicio.strftime("%d/%m/%Y"),
            "fin_fmt":      fecha_fin.strftime("%d/%m/%Y"),
            "val":          val_anual,
            "cuota_uf":     cuota_uf,
            "val_mes":      val_mes,
            "n_mant":       n_mant,
            "fact_flags":   fact_flags,   # internal, excluded from DATA output
            "real_ytd":     real_ytd,
            "real_2025":    real_2025,
            "real_2024":    real_2024,
            "dias_vence":   dias_vence,
            "long_dias":    long_dias,
            "tpo_activo":   tpo_activo,
            "pct_consumido": pct_consumido,
            "es_nuevo":     tpo_activo <= 90,
            "n_mant_actual":  n_mant_actual,
            "estado":         estado,
            "programa":       programa,
            "linea_negocio":  linea_negocio_contrato(num_str),
            "bajo_contrato":  bajo_contrato,
            "dias_inicio_cli": tpo_activo,
        })

    # Eliminar duplicados por número de contrato (filas ocultas/filtradas en Excel)
    seen = {}
    for c in contratos:
        key = c["n"]
        if key not in seen:
            seen[key] = c
        else:
            # Si el duplicado está Activado y el existente no, reemplazar
            if c["estado"] == "Activado" and seen[key]["estado"] != "Activado":
                seen[key] = c
    duplicados = len(contratos) - len(seen)
    if duplicados > 0:
        print(f"       ADVERTENCIA: {duplicados} contrato(s) duplicado(s) eliminado(s) de CONTRATOS TODOS")
    contratos_final = list(seen.values())

    # "es_nuevo" no debe depender sólo de la fecha de inicio del contrato: si
    # el cliente ya tenía OTRO contrato (cualquier estado) antes de éste, no es
    # un cliente nuevo, es una renovación/continuación bajo otro N° de contrato.
    primer_inicio_cliente = {}
    for c in contratos_final:
        cli = c["cliente"]
        if cli not in primer_inicio_cliente or c["inicio"] < primer_inicio_cliente[cli]:
            primer_inicio_cliente[cli] = c["inicio"]
    for c in contratos_final:
        c["es_nuevo"] = c["tpo_activo"] <= 90 and c["inicio"] == primer_inicio_cliente[c["cliente"]]

    # Un contrato "Activado" cuya fecha de término ya pasó, pero que tiene un
    # sucesor claro del mismo cliente (otro contrato que empieza dentro de una
    # ventana de 30 días desde ese término y termina más tarde), se considera
    # superado: el sucesor es "el que vale" y éste pasa a histórico (Estado ->
    # Expirado) para no contarlo en cartera activa, presupuesto ni Vencimientos.
    contratos_por_cliente_tmp = defaultdict(list)
    for c in contratos_final:
        contratos_por_cliente_tmp[c["cliente"]].append(c)
    superados = 0
    for c in contratos_final:
        if c["estado"] != "Activado":
            continue
        c_fin = date.fromisoformat(c["fin"])
        if c_fin >= TODAY:
            continue  # todavía no vence
        for otro in contratos_por_cliente_tmp[c["cliente"]]:
            if otro is c:
                continue
            o_fin = date.fromisoformat(otro["fin"])
            if o_fin <= c_fin:
                continue  # no extiende cobertura más allá de c
            o_inicio = date.fromisoformat(otro["inicio"])
            if abs((o_inicio - c_fin).days) <= 30:
                c["estado"] = "Expirado"
                superados += 1
                break
    if superados > 0:
        print(f"       {superados} contrato(s) vencido(s) marcados como histórico (tenían sucesor claro)")

    return contratos_final


# ══════════════════════════════════════════════════════════════════════════════
# 2. HOJA: FACTURACIÓN
#    Fila 1 = encabezados, Fila 2+ = datos
#    0=Cliente  1=NombreFact  2=Total2026  3=Total2025  4=Total2024
#    6=Contr2026  7=Contr2025  8=Contr2024
#    14=PrimeraVez  15=Renovo  16=NoContinuo  17=NContratos
# ══════════════════════════════════════════════════════════════════════════════
def read_facturacion(wb):
    ws = wb["FACTURACION"]
    panel = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        cliente = safe_str(row[0])
        if not cliente:
            continue

        # Col[2] = NombreAnalisis — nombre normalizado que coincide con BBDD col H
        nombre_analisis = safe_str(row[2]) or cliente

        # Col[3] = Total2026, Total2026 se desplazó a col[3]
        real_ytd   = to_float(row[3])
        real_2025  = to_float(row[4])
        real_2024  = to_float(row[5])
        contr_2026 = to_float(row[7])   # Columna H (Contr2026 YTD)
        contr_2025 = to_float(row[8])   # Columna I (Contratos 2025)
        contr_2024 = to_float(row[9])   # Columna J (Contratos 2024)
        # AZ-BK (idx 51-62): facturación contratos mes a mes 2026
        contr_meses = [to_float(row[51 + m]) if len(row) > 51 + m else 0.0 for m in range(12)]
        # BL (idx 63): flag "FACTURACIÓN < CONTRATOS" (SI/NO)
        fac_menor_contr = safe_str(row[63]).strip().upper() == "SI" if len(row) > 63 and row[63] else False
        n_contratos = to_int(row[18])   # desplazado por columna nueva

        primera_vez = safe_str(row[15]).upper() == "SI"  # desplazado
        renovo      = safe_str(row[16]).upper() == "SI"  # desplazado
        no_continuo = safe_str(row[17]).upper() == "SI"  # desplazado

        if primera_vez:
            estado_rel = "Nuevo"
        elif renovo:
            estado_rel = "Renovado"
        elif no_continuo:
            estado_rel = "Perdido"
        elif n_contratos > 0:
            estado_rel = "Con contrato"
        else:
            estado_rel = "Sin contrato"

        panel.append({
            "cliente":           cliente,
            "nombre_analisis":   nombre_analisis,   # key para lookup en BBDD
            "coord":             "Sin contrato",    # filled later from CONTRATOS
            "tipo_cli":          "Privado",          # filled later from BBDD
            "real_ytd":          real_ytd,
            "real_anual_2026":   real_ytd,           # YTD = mejor proxy para año en curso
            "real_anual_2025":   real_2025,
            "real_anual_2024":   real_2024,
            "real_ytd_2025":     0,                  # filled later from BBDD
            "real_ytd_2024":     0,
            "presup_contr_anio": 0,                 # filled later from CONTRATOS
            "presup_contr_ytd":  contr_2026,
            "contr_2025":        contr_2025,
            "contr_2024":        contr_2024,
            "contr_meses_2026":  contr_meses,
            "_fac_menor_contr":  fac_menor_contr,
            "n_contratos":       n_contratos,
            "tiene_contrato":    (n_contratos > 0) and not no_continuo,
            "estado_relacion":   estado_rel,
            "_no_continuo":      no_continuo,       # internal flag
        })

    return panel


# ══════════════════════════════════════════════════════════════════════════════
# 3. HOJA: BBDD FACTURACION (pandas — 32k+ filas)
#    Columnas clave (1-indexed en Excel → 0-indexed en pandas):
#    H(7)=NombreCliente  S(18)=TotalLineaVenta  AM(38)=Mes  AO(40)=Año
#    AR(43)=Empresa2  AT(45)=TipoCliente  AU(46)=LineaDeNegocio
# ══════════════════════════════════════════════════════════════════════════════
def read_bbdd(xlsx_path):
    print("  Leyendo BBDD FACTURACION con pandas (puede tardar ~15 s)...")
    df = pd.read_excel(xlsx_path, sheet_name="Facturacion a Fecha", header=0)

    cols = df.columns.tolist()
    # Access by position to avoid encoding issues in header names
    c_cliente   = cols[7]    # H
    c_monto     = cols[18]   # S: Total Linea Venta
    c_costo     = cols[16]   # Q: Total Costo Linea
    c_catalogo  = cols[36]   # AK: Catálogo
    c_tipodoc   = cols[37]   # AL: Tipo documento (Factura / Nota de crédito / etc.)
    c_mes       = cols[38]   # AM
    c_ano       = cols[40]   # AO
    c_emp2      = cols[43]   # AR: Empresa 2
    c_tipocli   = cols[45]   # AT: Tipo Cliente
    c_linea     = cols[46]   # AU: Linea de Negocio
    c_ejecutivo = cols[50] if len(cols) > 50 else None  # AY: Ejecutivo

    # Filter Tecservice rows
    df_ts = df[df[c_emp2].astype(str).str.strip() == "TS"].copy()

    df_ts[c_monto] = pd.to_numeric(df_ts[c_monto], errors="coerce").fillna(0)
    df_ts[c_costo] = pd.to_numeric(df_ts[c_costo], errors="coerce").fillna(0)
    df_ts[c_mes]   = pd.to_numeric(df_ts[c_mes],   errors="coerce")
    df_ts[c_ano]   = pd.to_numeric(df_ts[c_ano],   errors="coerce")
    df_ts = df_ts.dropna(subset=[c_mes, c_ano])
    df_ts[c_mes] = df_ts[c_mes].astype(int)
    df_ts[c_ano] = df_ts[c_ano].astype(int)

    # tipo_cli_map: last known sector per client
    tipo_map = (
        df_ts[df_ts[c_tipocli].notna()]
        .groupby(c_cliente)[c_tipocli]
        .last()
        .to_dict()
    )

    # Helper: build {year: {month: total}} dict from a sub-dataframe
    def monthly_dict(sub):
        result = {}
        for (ano, mes), grp in sub.groupby([c_ano, c_mes]):
            result.setdefault(int(ano), {})[int(mes)] = float(grp[c_monto].sum())
        return result

    mensual_total   = monthly_dict(df_ts)
    df_priv         = df_ts[df_ts[c_tipocli].astype(str).str.strip() == "Privado"]
    df_pub          = df_ts[df_ts[c_tipocli].astype(str).str.strip() == "Público"]
    df_nocontr      = df_ts[df_ts[c_linea].astype(str).str.strip() == "Ingresos No Recurrentes"]
    df_contr        = df_ts[df_ts[c_linea].astype(str).str.strip() != "Ingresos No Recurrentes"]

    mensual_priv    = monthly_dict(df_priv)
    mensual_pub     = monthly_dict(df_pub)
    mensual_contr   = monthly_dict(df_contr)
    mensual_nocontr = monthly_dict(df_nocontr)

    # Facturación mensual filtrada: solo Facturas + catálogos ST/Trazabilidad/REAS
    # Mismos filtros que el desglose por ejecutivo pero sin desagregar
    _CATALOGOS_FAC = {"Servicio Técnico", "Trazabilidad", "REAS"}
    df_ts[c_tipodoc]  = df_ts[c_tipodoc].astype(str).str.strip()
    df_ts[c_catalogo] = df_ts[c_catalogo].astype(str).str.strip()
    df_facturado = df_ts[
        (df_ts[c_tipodoc]  == "Factura") &
        (df_ts[c_catalogo].isin(_CATALOGOS_FAC))
    ].copy()
    mensual_facturado = monthly_dict(df_facturado)

    # Desglose "Otros Ingresos" del año actual: Trazabilidad / REAS / Marca (dentro de Servicio Técnico)
    # Usado por la sección "Desglose de Ingresos" (ver compute_desglose_ingresos)
    c_marca = cols[11]  # L: Marca
    df_fac_ano = df_facturado[df_facturado[c_ano] == ANO]

    def monthly_arr_12(sub):
        arr = [0.0] * 12
        for mes, grp in sub.groupby(c_mes):
            if 1 <= mes <= 12:
                arr[mes - 1] = float(grp[c_monto].sum())
        return arr

    traz_mensual = monthly_arr_12(df_fac_ano[df_fac_ano[c_catalogo] == "Trazabilidad"])
    reas_mensual = monthly_arr_12(df_fac_ano[df_fac_ano[c_catalogo] == "REAS"])

    df_st_ano = df_fac_ano[df_fac_ano[c_catalogo] == "Servicio Técnico"].copy()
    df_st_ano[c_marca] = df_st_ano[c_marca].astype(str).str.strip()
    marca_mensual = {
        marca: monthly_arr_12(grp)
        for marca, grp in df_st_ano.groupby(c_marca)
    }
    st_total_mensual = monthly_arr_12(df_st_ano)

    # Facturación mensual por ejecutivo (columna AY)
    # Usa df_facturado (ya filtrado) restringido al año actual
    mensual_por_ejecutivo = {}
    if c_ejecutivo:
        df_ts[c_ejecutivo] = df_ts[c_ejecutivo].astype(str).str.strip()
        df_eje_base = df_facturado[df_facturado[c_ano] == ANO].copy()
        ejecutivos = [e for e in df_eje_base[c_ejecutivo].unique()
                      if isinstance(e, str) and e.strip().lower() not in ("nan", "none", "")]
        for eje in ejecutivos:
            df_eje = df_eje_base[df_eje_base[c_ejecutivo] == eje]
            mensual_por_ejecutivo[eje] = monthly_dict(df_eje)

    # MES_CORTE: mes calendario actual (dinámico)
    mes_corte = TODAY.month

    # YTD per client: solo Facturas + catálogos ST/REAS/Trazabilidad
    _CATALOGOS_YTD = {"Servicio Técnico", "REAS", "Trazabilidad"}
    df_ytd_base = df_ts[
        (df_ts[c_tipodoc] == "Factura") &
        df_ts[c_catalogo].isin(_CATALOGOS_YTD)
    ]

    def ytd_per_cli(year, max_mes):
        sub = df_ytd_base[(df_ytd_base[c_ano] == year) & (df_ytd_base[c_mes] <= max_mes)]
        return sub.groupby(c_cliente)[c_monto].sum().to_dict()

    ytd_cli_2025 = ytd_per_cli(ANO - 1, mes_corte)
    ytd_cli_2024 = ytd_per_cli(ANO - 2, mes_corte)
    # Año en curso con la MISMA base (TS + Factura + los 3 catálogos de
    # servicio). Es la que cuadra con "Ingresos Totales" de la hoja Analisis
    # Facturación: excluye Provisión y los catálogos de venta de equipos
    # (E./C. Esterilización), que no son ingreso de servicio.
    ytd_cli_2026 = ytd_per_cli(ANO, mes_corte)
    # Costo del año en curso por cliente, con la misma base
    _sub26 = df_ytd_base[(df_ytd_base[c_ano] == ANO) & (df_ytd_base[c_mes] <= mes_corte)]
    costo_cli_2026 = _sub26.groupby(c_cliente)[c_costo].sum().to_dict()
    ytd_2026_tot = float(
        df_ytd_base[(df_ytd_base[c_ano] == ANO) & (df_ytd_base[c_mes] <= mes_corte)][c_monto].sum()
    )

    # YTD contratos 2024/2025: Vendedor (col T) empieza con "ST"
    c_vendedor = cols[19]
    df_ytd_base = df_ytd_base.copy()
    df_ytd_base[c_vendedor] = df_ytd_base[c_vendedor].astype(str).str.strip().str.upper()
    df_ytd_contr = df_ytd_base[df_ytd_base[c_vendedor].str.startswith("ST")]

    def ytd_contr_total(year, max_mes):
        sub = df_ytd_contr[(df_ytd_contr[c_ano] == year) & (df_ytd_contr[c_mes] <= max_mes)]
        return float(sub[c_monto].sum())

    ytd_contr_2024 = ytd_contr_total(ANO - 2, mes_corte)
    ytd_contr_2025 = ytd_contr_total(ANO - 1, mes_corte)

    # Monthly billing per client for current year (same filters as df_fac_ano)
    mensual_por_cliente = {}
    for cli_k, grp in df_fac_ano.groupby(c_cliente):
        s = str(cli_k).strip()
        if not s or s.lower() in ("nan", "none"):
            continue
        a = [0.0] * 12
        for m_val, sub in grp.groupby(c_mes):
            mi = int(m_val) - 1
            if 0 <= mi < 12:
                a[mi] = float(sub[c_monto].sum())
        mensual_por_cliente[s] = a

    return {
        "mensual_total":         mensual_total,
        "mensual_facturado":     mensual_facturado,
        "traz_mensual":          traz_mensual,
        "reas_mensual":          reas_mensual,
        "marca_mensual":         marca_mensual,
        "st_total_mensual":      st_total_mensual,
        "mensual_priv":          mensual_priv,
        "mensual_pub":           mensual_pub,
        "mensual_contr":         mensual_contr,
        "mensual_nocontr":       mensual_nocontr,
        "mensual_por_ejecutivo": mensual_por_ejecutivo,
        "tipo_cli_map":          tipo_map,
        "ytd_cli_2025":          ytd_cli_2025,
        "ytd_cli_2024":          ytd_cli_2024,
        "ytd_cli_2026":          ytd_cli_2026,
        "costo_cli_2026":        costo_cli_2026,
        "ytd_2026_tot":          ytd_2026_tot,
        "ytd_contr_2024":        ytd_contr_2024,
        "ytd_contr_2025":        ytd_contr_2025,
        "mes_corte":             mes_corte,
        "mensual_por_cliente":   mensual_por_cliente,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 4. HOJA: VISITAS
#    0=Asignado  2=Fecha  3=Cliente  6=TipoActividad  8=Conteo  9=Mes  10=Año
# ══════════════════════════════════════════════════════════════════════════════
def read_visitas(wb, mes_corte):
    ws = wb["VISITAS"]

    mensual        = {e: {str(ANO - 1): [0] * 12, str(ANO): [0] * 12} for e in EJECUTIVOS_VISITAS}
    tipo_acts      = {e: defaultdict(int) for e in EJECUTIVOS_VISITAS}
    cli_count_ytd  = {e: defaultdict(int) for e in EJECUTIVOS_VISITAS}
    cli_count_all  = {e: defaultdict(int) for e in EJECUTIVOS_VISITAS}
    unique_2026    = {e: set() for e in EJECUTIVOS_VISITAS}
    unique_2025_ytd = {e: set() for e in EJECUTIVOS_VISITAS}
    # Visitas mensuales por ejecutivo + cliente (para gráfico de búsqueda)
    cli_mensual    = {e: defaultdict(lambda: {str(ANO-1): [0]*12, str(ANO): [0]*12})
                      for e in EJECUTIVOS_VISITAS}

    for row in ws.iter_rows(min_row=2, values_only=True):
        asignado = safe_str(row[0])
        if asignado not in EJECUTIVOS_VISITAS:
            continue

        tipo_act = safe_str(row[6])
        cliente  = safe_str(row[3])
        conteo   = to_int(row[8], 1)
        mes      = to_int(row[9], 0)
        ano      = to_int(row[10], 0)

        if not (1 <= mes <= 12) or not ano:
            continue

        ano_str = str(ano)
        if ano_str in mensual[asignado]:
            mensual[asignado][ano_str][mes - 1] += conteo
            cli_mensual[asignado][cliente][ano_str][mes - 1] += conteo

        tipo_acts[asignado][tipo_act] += conteo
        cli_count_all[asignado][cliente] += conteo

        if ano == ANO and mes <= mes_corte:
            cli_count_ytd[asignado][cliente] += conteo
            unique_2026[asignado].add(cliente)

        if ano == ANO - 1 and mes <= mes_corte:
            unique_2025_ytd[asignado].add(cliente)

    # Build resumen
    resumen = {}
    for e in EJECUTIVOS_VISITAS:
        tot_2025 = sum(mensual[e][str(ANO - 1)])
        tot_2026 = sum(mensual[e][str(ANO)])
        ytd_2026 = sum(mensual[e][str(ANO)][:mes_corte])
        ytd_2025_mismo = sum(mensual[e][str(ANO - 1)][:mes_corte])
        resumen[e] = {
            "total":              tot_2025 + tot_2026,
            "tot_2025":           tot_2025,
            "tot_2026":           tot_2026,
            "tot_2026_ytd":       ytd_2026,
            "tot_2025_ytd_mismo": ytd_2025_mismo,
            "clientes_unicos_2026": len(unique_2026[e]),
        }

    # Top 10 clients by YTD visits
    top = {}
    for e in EJECUTIVOS_VISITAS:
        sorted_cli = sorted(cli_count_ytd[e].items(), key=lambda x: -x[1])[:10]
        top[e] = [{"cliente": c, "n": n} for c, n in sorted_cli]

    # Convertir cli_mensual (defaultdict anidado) a dict serializable
    cli_mensual_out = {}
    for e in EJECUTIVOS_VISITAS:
        cli_mensual_out[e] = {cli: dict(meses) for cli, meses in cli_mensual[e].items()}

    return {
        "mensual":     {e: dict(mensual[e]) for e in EJECUTIVOS_VISITAS},
        "tipo":        {e: dict(tipo_acts[e]) for e in EJECUTIVOS_VISITAS},
        "resumen":     resumen,
        "top":         top,
        "cli_mensual": cli_mensual_out,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 5. HOJA: SATISFACCION
#    0=Version  6=Estado  7=Nombre  8=Correo  9=Institucion
#   10=Calidad(0-7)  11=Resuelto  12=Tiempo(0-7)  13=Recomendacion(0-10)
#   14=Comentario  15=NombreAnalisis  16=NombreBI  17=BITotal
# ══════════════════════════════════════════════════════════════════════════════

# Correcciones de nombres con errores tipográficos o versiones abreviadas
_NOMBRE_FIXES = {
    "CLINIDA VAILA VESPUCIO":         "CLINICA DAVILA VESPUCIO",
    "HOSPITAL BARROS LUCO":           "HOSPITAL BARROS LUCO TRUDEAU",
    "HOSPITAL SAN LUIS DE BUIN":      "HOSPITAL SAN LUIS DE BUIN PAINE",
    "HOSPITAL SAN LUIS DE BUIN PAINE.": "HOSPITAL SAN LUIS DE BUIN PAINE",
}

_PERSONAL_DOMAINS = {
    "gmail.com", "hotmail.com", "yahoo.com", "outlook.com", "live.com",
    "icloud.com", "yahoo.es", "hotmail.es", "protonmail.com", "desconocido",
}

def _dom_cat(dom):
    if "gob.cl" in dom or ".gob." in dom:
        return "Sector Público"
    if dom in _PERSONAL_DOMAINS:
        return "Correo personal"
    return "Sector Privado"


def read_satisfaccion(wb):
    # Leer mapa de nombres normalizados
    ws_map = wb["Nombre Satisfacción"]
    nombre_map = {}
    for row in ws_map.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            nombre_map[safe_str(row[0]).lower()] = safe_str(row[1]).upper()

    ws = wb["SATISFACCION"]
    respuestas = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        estado = safe_str(row[6]).lower()
        if estado != "completed":
            continue

        inst_raw        = safe_str(row[9])
        nombre_analisis = safe_str(row[15])
        if not nombre_analisis:
            nombre_analisis = nombre_map.get(inst_raw.lower(), inst_raw.upper())

        # Corregir errores tipográficos y nombres abreviados
        nombre_analisis = _NOMBRE_FIXES.get(nombre_analisis, nombre_analisis)

        # col[16] = NombreBI: nombre exacto para cruzar con Base Instalada
        nombre_bi = safe_str(row[16]).strip().upper() if len(row) > 16 and row[16] else ""

        email   = safe_str(row[8])
        dominio = email.split("@")[1].lower() if "@" in email else "desconocido"

        calidad  = to_float(row[10])
        resuelto = safe_str(row[11])
        tiempo   = to_float(row[12])
        recom    = to_float(row[13])
        mejora   = safe_str(row[14])
        # Columna AD (row[29]) = categoría detractor (llenada manualmente en Excel)
        cat_det  = safe_str(row[29]).strip() if len(row) > 29 and row[29] else ""
        # Columna R (row[17]) = contratos asociados al cliente (monto CLP)
        contr_asoc = round(to_float(row[17] if len(row) > 17 else 0))
        _BD_KEYS = ["ester","endo","mob","dent","inc","mmq","reas","otro"]
        bi_det   = {k: to_int(row[18+i] if len(row) > 18+i else None)
                    for i, k in enumerate(_BD_KEYS)}
        # bi_total = suma real de equipos por tipo (no la columna R)
        bi_sum   = sum(bi_det.values())
        bi       = bi_sum if bi_sum > 0 else None
        # Columna AC (row[28]) = facturación real 2026 del cliente
        fac_2026_sat = round(to_float(row[28] if len(row) > 28 else 0))
        # Columna E (row[4]) = Fecha y hora de finalización de la encuesta
        fecha_resp = row[4] if len(row) > 4 and isinstance(row[4], datetime) else None

        respuestas.append({
            "institucion":    nombre_analisis,
            "nombre_bi":      nombre_bi,
            "dominio":        dominio,
            "categoria":      _dom_cat(dominio),
            "fecha":          fecha_resp,
            "calidad":        calidad,
            "tiempo":         tiempo,
            "recom":          recom,
            "resuelto":       resuelto,
            "mejora":         mejora,
            "cat_detractor":  cat_det,
            "bi_total":       bi,
            "bi_detalle":     bi_det,
            "contr_asociados": contr_asoc,
            "fac_2026":       fac_2026_sat,
        })

    n = len(respuestas)
    if n == 0:
        return {
            "global":        {"n": 0, "calidad_avg": 0, "tiempo_avg": 0, "recom_avg": 0,
                              "resuelto_si": 0, "resuelto_no": 0, "resuelto_parcial": 0},
            "nps":           {"det": 0, "pas": 0, "pro": 0, "nps": 0},
            "dominio":       [],
            "categoria":     [],
            "trimestral":            [],
            "detractor_categorias":  [],
            "instituciones": [],
            "bi_resumen":    {"n_inst": 0, "n_con_bi": 0, "total_bi": 0},
            "comentarios":   [],
        }

    cal_avg = round(sum(r["calidad"] for r in respuestas) / n, 2)
    tie_avg = round(sum(r["tiempo"]  for r in respuestas) / n, 2)
    rec_avg = round(sum(r["recom"]   for r in respuestas) / n, 2)

    def _resuelto_cat(s):
        sl = s.lower()
        if "parcial" in sl: return "parcial"
        if "no" in sl and "si" not in sl: return "no"
        return "si"

    res_si  = sum(1 for r in respuestas if _resuelto_cat(r["resuelto"]) == "si")
    res_no  = sum(1 for r in respuestas if _resuelto_cat(r["resuelto"]) == "no")
    res_par = n - res_si - res_no

    # NPS
    pro = sum(1 for r in respuestas if r["recom"] >= 9)
    pas = sum(1 for r in respuestas if 7 <= r["recom"] <= 8)
    det = sum(1 for r in respuestas if r["recom"] <= 6)
    nps = round((pro - det) / n * 100)

    # Por dominio
    dom_data = defaultdict(lambda: {"n": 0, "cal": [], "tie": [], "rec": []})
    for r in respuestas:
        d = dom_data[r["dominio"]]
        d["n"] += 1
        d["cal"].append(r["calidad"])
        d["tie"].append(r["tiempo"])
        d["rec"].append(r["recom"])

    dominio_list = []
    for dom, d in sorted(dom_data.items(), key=lambda x: -x[1]["n"]):
        dominio_list.append({
            "dominio":   dom,
            "n":         d["n"],
            "calidad":   round(sum(d["cal"]) / d["n"], 2),
            "tiempo":    round(sum(d["tie"]) / d["n"], 2),
            "recom":     round(sum(d["rec"]) / d["n"], 2),
            "categoria": _dom_cat(dom),
        })

    # Por categoría (ponderado)
    cat_data: dict = {}
    for entry in dominio_list:
        cat = entry["categoria"]
        if cat not in cat_data:
            cat_data[cat] = {"n": 0, "cal": [], "tie": [], "rec": []}
        cat_data[cat]["n"] += entry["n"]
        cat_data[cat]["cal"].append((entry["calidad"], entry["n"]))
        cat_data[cat]["tie"].append((entry["tiempo"],  entry["n"]))
        cat_data[cat]["rec"].append((entry["recom"],   entry["n"]))

    def w_avg(pairs):
        total_n = sum(p[1] for p in pairs)
        return 0 if total_n == 0 else round(sum(p[0] * p[1] for p in pairs) / total_n, 2)

    categoria_list = [
        {"categoria": cat, "n": d["n"],
         "calidad": w_avg(d["cal"]), "tiempo": w_avg(d["tie"]), "recom": w_avg(d["rec"])}
        for cat, d in sorted(cat_data.items(), key=lambda x: -x[1]["n"])
    ]

    # Evolutivo trimestral: N° de encuestas y satisfacción promedio por trimestre
    trim_data: dict = {}
    for r in respuestas:
        f = r.get("fecha")
        if f is None:
            continue
        q = (f.month - 1) // 3 + 1
        key = (f.year, q)
        if key not in trim_data:
            trim_data[key] = {"cal": [], "tie": [], "rec": [], "pro": 0, "pas": 0, "det": 0}
        t = trim_data[key]
        t["cal"].append(r["calidad"])
        t["tie"].append(r["tiempo"])
        t["rec"].append(r["recom"])
        if r["recom"] >= 9: t["pro"] += 1
        elif r["recom"] >= 7: t["pas"] += 1
        else: t["det"] += 1

    trimestral_list = [
        {
            "trimestre": f"Q{q} {year}",
            "n":         len(t["cal"]),
            "calidad":   round(sum(t["cal"]) / len(t["cal"]), 2),
            "tiempo":    round(sum(t["tie"]) / len(t["tie"]), 2),
            "recom":     round(sum(t["rec"]) / len(t["rec"]), 2),
            "promotores":  t["pro"],
            "pasivos":     t["pas"],
            "detractores": t["det"],
            "nps": round((t["pro"] - t["det"]) / len(t["cal"]) * 100) if t["cal"] else 0,
        }
        for (year, q), t in sorted(trim_data.items())
    ]

    # Desglose por categoría detractor (columna "Categoría" del Excel, se llena
    # manualmente sólo para respuestas con problemas). Sólo categorías con valor.
    det_cat_data: dict = {}
    n_con_categoria = 0
    for r in respuestas:
        c = (r.get("cat_detractor") or "").strip()
        if not c:
            continue
        n_con_categoria += 1
        if c not in det_cat_data:
            det_cat_data[c] = {"n": 0, "rec": []}
        det_cat_data[c]["n"] += 1
        det_cat_data[c]["rec"].append(r["recom"])

    detractor_categorias = [
        {"categoria": c, "n": d["n"], "recom_avg": round(sum(d["rec"]) / d["n"], 2)}
        for c, d in sorted(det_cat_data.items(), key=lambda x: -x[1]["n"])
    ]

    # Por institución (TODAS las respuestas, no solo las que tienen mejora)
    # BI = primer valor no-nulo (mismo cliente → mismo valor)
    # Categoría = preferir Sector Público o Privado sobre Correo personal
    _CAT_PRIO  = {"Sector Público": 0, "Sector Privado": 1, "Correo personal": 2}
    _BD_KEYS   = ["ester","endo","mob","dent","inc","mmq","reas","otro"]
    _BD_EMPTY  = {k: 0 for k in _BD_KEYS}
    inst_data: dict = {}
    for r in respuestas:
        key = r["institucion"]
        if key not in inst_data:
            inst_data[key] = {
                "n": 0, "cal": [], "tie": [], "rec": [],
                "_bi": None, "_bi_det": None, "_contr": None, "_fac2026": None,
                "_nombre_bi": r.get("nombre_bi", ""),
                "_cat_det": r.get("cat_detractor", ""),
                "categoria": r["categoria"]
            }
        d = inst_data[key]
        d["n"] += 1
        d["cal"].append(r["calidad"])
        d["tie"].append(r["tiempo"])
        d["rec"].append(r["recom"])
        if d["_bi"] is None and r["bi_total"] is not None:
            d["_bi"] = r["bi_total"]
        # bi_detalle: tomar primera fila con suma > 0 (todos son el mismo cliente)
        if d["_bi_det"] is None:
            bd = r.get("bi_detalle", {})
            if sum(bd.values()) > 0:
                d["_bi_det"] = bd
        # contr_asociados: primer valor no-nulo (mismo cliente = mismo monto siempre)
        if d["_contr"] is None and r.get("contr_asociados", 0) > 0:
            d["_contr"] = r["contr_asociados"]
        # fac_2026: primer valor no-nulo de columna AC
        if d["_fac2026"] is None and r.get("fac_2026", 0) > 0:
            d["_fac2026"] = r["fac_2026"]
        # nombre_bi: primer valor no vacío (mismo cliente = mismo valor)
        if not d["_nombre_bi"] and r.get("nombre_bi"):
            d["_nombre_bi"] = r["nombre_bi"]
        # cat_detractor: primer valor no vacío entre todas las respuestas de la institución
        if not d["_cat_det"] and r.get("cat_detractor"):
            d["_cat_det"] = r["cat_detractor"]
        if _CAT_PRIO.get(r["categoria"], 9) < _CAT_PRIO.get(d["categoria"], 9):
            d["categoria"] = r["categoria"]

    instituciones = sorted([
        {
            "institucion":    key,
            "n":              d["n"],
            "calidad":        round(sum(d["cal"]) / d["n"], 2),
            "tiempo":         round(sum(d["tie"]) / d["n"], 2),
            "recom":          round(sum(d["rec"]) / d["n"], 2),
            "bi_total":       d["_bi"],
            "bi_detalle":     d["_bi_det"] or dict(_BD_EMPTY),
            "contr_asociados": d["_contr"] or 0,
            "fac_2026":       d["_fac2026"] or 0,
            "nombre_bi":      d["_nombre_bi"] or "",
            "cat_detractor":  d["_cat_det"] or "",
            "categoria":      d["categoria"],
        }
        for key, d in inst_data.items()
    ], key=lambda x: (-(x["bi_total"] or 0), x["institucion"]))

    inst_con_bi = [i for i in instituciones if i["bi_total"] is not None]
    bi_resumen = {
        "n_inst":    len(instituciones),
        "n_con_bi":  len(inst_con_bi),
        "total_bi":  sum(i["bi_total"] for i in inst_con_bi),
    }

    comentarios = [
        {"institucion": r["institucion"], "mejora": r["mejora"], "recom": r["recom"],
         "bi_total": r["bi_total"], "contr_asociados": r.get("contr_asociados", 0)}
        for r in respuestas if r["mejora"]
    ]

    return {
        "global": {
            "n":               n,
            "calidad_avg":     cal_avg,
            "tiempo_avg":      tie_avg,
            "recom_avg":       rec_avg,
            "resuelto_si":     res_si,
            "resuelto_no":     res_no,
            "resuelto_parcial": res_par,
        },
        "nps":           {"det": det, "pas": pas, "pro": pro, "nps": nps},
        "dominio":       dominio_list,
        "categoria":     categoria_list,
        "trimestral":            trimestral_list,
        "detractor_categorias":  detractor_categorias,
        "instituciones": instituciones,
        "bi_resumen":    bi_resumen,
        "comentarios":   comentarios,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 5b. HOJA: BASE INSTALADA
#     col[0]=Habilitado(Si/No)  col[1]=FueraDeServicio(Si/No)
#     col[4]=Nombre  col[5]=Fabricante  col[6]=Modelo
#     col[12]=Tipo  col[13]=Clasificacion1(cliente)  col[14]=Clasificacion2(estado)
#     col[25]=LineaDeNegocio  col[26]=NombreAnalisis(cliente normalizado, puede ser None)
#     col[27]=PotencialST (Si/No) — marcas que TECSERVICE aún representa
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# PROSPECTOS BI — la base instalada leída como cartera de renovación
# ══════════════════════════════════════════════════════════════════════════════
# Mismas filas activas que read_base_instalada(), pero conservando las dos
# columnas que esa función ignora y que son las que permiten prospectar:
#   col Q  (16) «Fecha de Compra»    → la instalación del equipo, de donde sale
#                                       su vida (hoy − fecha)
#   col AF (31) «Valorización CLP»   → cuánto vale reponerlo
#
# Se emite a nivel de equipo y no agregado: la prospección se hace cruzando
# libremente región, línea, tipo, estado y un umbral de vida que el usuario
# mueve en pantalla, y cualquier pre-agregación cerraría esas combinaciones.
# Son ~13.000 filas de enteros, que comprimen bien.
#
# La fecha se guarda como un solo entero ym = año*12 + (mes−1), y −1 cuando la
# fila no la trae. Sólo un tercio de la base tiene fecha y un sexto tiene
# valorización: el panel muestra esa cobertura en vez de disimularla, porque
# un potencial calculado sobre datos incompletos hay que leerlo sabiéndolo.

def read_prospectos_bi(wb):
    """Base instalada abierta por equipo, con fecha de instalación y valor."""
    ws = None
    for name in wb.sheetnames:
        if "base instalada" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        return {}

    _EST = {
        "CONTRATO": "Contrato", "CONTRATO 24/7": "Contrato",
        "GARANTÍA": "Garantía", "GARANTIA": "Garantía",
        "SIN GARANTIA": "Sin garantía",
    }

    regiones, lineas, tipos, clientes, estados = [], [], [], [], []
    i_reg, i_lin, i_tip, i_cli, i_est = {}, {}, {}, {}, {}
    # Identificación del equipo, para poder mostrar el detalle bajo cada
    # cliente: nombre (col E), fabricante (F), modelo (G) y serie (H).
    nombres, fabricantes, modelos, series = [], [], [], []
    i_nom, i_fab, i_mod, i_ser = {}, {}, {}, {}

    def ref(lista, indice, valor):
        if valor not in indice:
            indice[valor] = len(lista)
            lista.append(valor)
        return indice[valor]

    filas_out = []
    con_fecha = con_valor = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if safe_str(row[0]).upper() != "SI":
            continue
        if safe_str(row[1]).upper() == "SI":
            continue

        tipo = (safe_str(row[12]).strip().upper() or "SIN TIPO") if row[12] else "SIN TIPO"
        linea = safe_str(row[25]).strip() if row[25] else "Otra"
        estado = _EST.get(safe_str(row[14]).strip().upper() if row[14] else "", "Sin clasificar")
        region = _norm_region(row[29] if len(row) > 29 else None)

        cli = safe_str(row[26]).strip() if row[26] else ""
        if not cli or cli.lower() in ("none", ""):
            cli = safe_str(row[13]).strip() if row[13] else "SIN CLIENTE"
        cli = cli.upper()
        # GEMCO es equipamiento propio, no cartera que prospectar. Se excluye
        # igual que en read_base_instalada() para que los totales de las dos
        # hojas hablen del mismo universo.
        if "GEMCO" in cli:
            continue

        pot = safe_str(row[27]).strip().upper() if len(row) > 27 and row[27] is not None else ""
        es_pot = 1 if pot in ("SI", "SÍ", "S", "1", "TRUE", "VERDADERO") else 0

        f = row[16] if len(row) > 16 else None
        if isinstance(f, (datetime, date)):
            ym = f.year * 12 + (f.month - 1)
            dia = f.day
            con_fecha += 1
        else:
            ym, dia = -1, 0

        v = row[31] if len(row) > 31 else None
        valor = int(round(to_float(v))) if isinstance(v, (int, float)) and v else 0
        if valor:
            con_valor += 1

        filas_out.append([
            ref(regiones, i_reg, region),
            ref(lineas, i_lin, linea),
            ref(tipos, i_tip, tipo),
            ref(clientes, i_cli, cli),
            ref(estados, i_est, estado),
            es_pot, ym, valor,
            ref(nombres, i_nom, safe_str(row[4]).strip() if row[4] else ""),
            ref(fabricantes, i_fab, safe_str(row[5]).strip() if row[5] else ""),
            ref(modelos, i_mod, safe_str(row[6]).strip() if row[6] else ""),
            ref(series, i_ser, safe_str(row[7]).strip() if row[7] else ""),
            dia,
        ])

    hoy = TODAY
    return {
        "regiones": regiones,
        "lineas":   lineas,
        "tipos":    tipos,
        "clientes": clientes,
        "estados":  estados,
        "nombres":     nombres,
        "fabricantes": fabricantes,
        "modelos":     modelos,
        "series":      series,
        "hoy_ym":   hoy.year * 12 + (hoy.month - 1),
        "vida_util": 10,          # años de referencia para la renovación
        "n":         len(filas_out),
        "con_fecha": con_fecha,
        "con_valor": con_valor,
        # [región, línea, tipo, cliente, estado, potencialST, ym, valor,
        #  nombre, fabricante, modelo, serie, día]
        "filas":     filas_out,
    }


# ══════════════════════════════════════════════════════════════════════════════
# DESGLOSE DE FACTURACIÓN POR CLIENTE
# ══════════════════════════════════════════════════════════════════════════════
# La hoja FACTURACIÓN ya trae, cliente por cliente, la facturación total del año
# separada entre lo que viene de contratos y lo que no ("Correctiva"). Es el
# único lugar donde ese corte existe a nivel de cliente: el resto del panel sólo
# sabe el total facturado y si el cliente tiene contrato o no.
#
# Las dos partes suman exactamente el total, incluso cuando "Correctiva" sale
# negativa: eso ocurre cuando lo devengado del contrato supera lo efectivamente
# facturado en el año, y es información, no un error que haya que recortar.

def read_fact_desglose(wb):
    """Facturación por cliente abierta en contratos y no contratos."""
    ws = None
    for name in wb.sheetnames:
        if name.strip().upper() == "FACTURACION":
            ws = wb[name]
            break
    if ws is None:
        print("  ADVERTENCIA: no se encontro la hoja FACTURACION.")
        return {}

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return {}
    cab = filas[0]

    def col(*frag):
        for f in frag:
            fn = _norm_cli(f)
            for i, v in enumerate(cab):
                if v is not None and fn in _norm_cli(v):
                    return i
        return None

    C = {
        "nombre": col("NOMBRA ANALISIS", "NOMBRE ANALISIS", "NOMBRE DEL CLIENTE"),
        "t26":    col("FACTURACION TOTAL 2026"),
        "t25":    col("FACTURACION TOTAL 2025"),
        "c26":    col("FACTURACION TOTAL CONTRATOS 2026"),
        "c25":    col("FACTURACION TOTAL CONTRATOS 2025"),
        "k26":    col("FACTURACION TOTAL CORRECTIVA 2026"),
        "k25":    col("FACTURACION TOTAL CORRECTIVA 2025"),
        "bi":     col("BI TOTAL"),
        "tipo":   col("TIPO DE CLIENTE"),
        "nctr":   col("RECUENTO NUMERO CONTRATOS"),
    }
    faltan = [k for k, v in C.items() if v is None]
    if faltan:
        print(f"  ADVERTENCIA: columnas no encontradas en FACTURACION: {', '.join(faltan)}")

    def num(row, i):
        if i is None or i >= len(row):
            return 0.0
        v = row[i]
        return to_float(v) if isinstance(v, (int, float)) else 0.0

    out = {}
    for row in filas[1:]:
        if not row:
            continue
        nom = safe_str(row[C["nombre"]]).strip() if C["nombre"] is not None else ""
        if not nom:
            continue
        k = _norm_cli(nom)
        d = out.get(k)
        if d is None:
            d = out[k] = {"nombre": nom, "t26": 0.0, "t25": 0.0, "c26": 0.0,
                          "c25": 0.0, "k26": 0.0, "k25": 0.0, "bi": 0,
                          "tipo": "", "n_contratos": 0}
        # Un mismo cliente puede venir en más de una fila (alias de facturación):
        # se suman los montos y se conserva el mayor recuento de contratos.
        for f in ("t26", "t25", "c26", "c25", "k26", "k25"):
            d[f] += num(row, C[f])
        d["bi"] = max(d["bi"], int(num(row, C["bi"])))
        d["n_contratos"] = max(d["n_contratos"], int(num(row, C["nctr"])))
        if not d["tipo"] and C["tipo"] is not None and C["tipo"] < len(row):
            d["tipo"] = safe_str(row[C["tipo"]]).strip()

    print(f"       Desglose facturacion: {len(out)} clientes")
    return out


def read_base_instalada(wb):
    ws = None
    for name in wb.sheetnames:
        if "base instalada" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        print("  ADVERTENCIA: no se encontro la hoja 'BASE INSTALADA'.")
        return {
            "total": 0, "por_linea": {}, "por_tipo": [],
            "por_estado": {}, "clientes": [],
        }

    _ESTADO_MAP = {
        "CONTRATO":     "Contrato",
        "CONTRATO 24/7":"Contrato",
        "GARANTÍA":     "Garantia",
        "GARANTIA":     "Garantia",
        "SIN GARANTIA": "Sin garantia",
    }

    por_linea      = defaultdict(int)
    por_tipo       = defaultdict(int)
    por_tipo_si    = defaultdict(int)   # solo filas con Potencial ST = Si
    por_tipo_no    = defaultdict(int)   # solo filas con Potencial ST = No
    por_estado     = defaultdict(int)
    tipo_por_linea = defaultdict(lambda: defaultdict(int))
    cli_map        = {}  # cliente → {total, lineas:{}, estados:{}}
    # Región: viene en la col AD de la propia hoja, llena en el 100% de las
    # filas activas. No se cruza contra BASE MAPA porque ésa sólo cubre los
    # clientes con facturación (182 de 1.501).
    reg_map        = {}

    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Filtro activos
        habilitado = safe_str(row[0]).upper() if row[0] is not None else ""
        fuera_sv   = safe_str(row[1]).upper() if row[1] is not None else ""
        if habilitado != "SI":
            continue
        if fuera_sv == "SI":
            continue

        tipo   = safe_str(row[12]).strip().upper() if row[12] else "SIN TIPO"
        linea  = safe_str(row[25]).strip() if row[25] else "Otra"
        estado_raw = safe_str(row[14]).strip().upper() if row[14] else ""
        estado = _ESTADO_MAP.get(estado_raw, "Sin clasificar")

        # Nombre cliente: col[26] si existe, sino col[13]
        nombre_analisis = safe_str(row[26]).strip() if row[26] else ""
        if not nombre_analisis or nombre_analisis.lower() in ("none", ""):
            nombre_analisis = safe_str(row[13]).strip() if row[13] else "SIN CLIENTE"
        nombre_analisis = nombre_analisis.upper()

        # col[27] = Potencial ST (Si/No)
        potencial_raw = safe_str(row[27]).strip().upper() if len(row) > 27 and row[27] is not None else ""
        es_potencial  = potencial_raw in ("SI", "SÍ", "S", "1", "TRUE", "VERDADERO")

        # col[29] = Región (col AD)
        region = _norm_region(row[29] if len(row) > 29 else None)
        if "GEMCO" not in nombre_analisis:
            rd = reg_map.setdefault(region, {
                "total": 0, "total_si": 0,
                "lineas": defaultdict(int), "lineas_si": defaultdict(int),
                "_clientes": set(),
            })
            rd["total"] += 1
            rd["lineas"][linea] += 1
            rd["_clientes"].add(nombre_analisis)
            if es_potencial:
                rd["total_si"] += 1
                rd["lineas_si"][linea] += 1

        total += 1
        por_linea[linea] += 1
        por_tipo[tipo]   += 1
        if es_potencial: por_tipo_si[tipo] += 1
        else:            por_tipo_no[tipo] += 1
        por_estado[estado] += 1
        tipo_por_linea[linea][tipo] += 1

        if nombre_analisis not in cli_map:
            cli_map[nombre_analisis] = {
                "total": 0, "total_si": 0, "total_no": 0,
                "lineas": defaultdict(int),
                "lineas_si": defaultdict(int),
                "lineas_no": defaultdict(int),
                "estados": defaultdict(int),
                "regiones": defaultdict(int),
                "_potencial_st": False,
            }
        d = cli_map[nombre_analisis]
        d["total"]          += 1
        d["lineas"][linea]  += 1
        d["estados"][estado] += 1
        d["regiones"][region] += 1
        if es_potencial:
            d["_potencial_st"]     = True
            d["total_si"]          += 1
            d["lineas_si"][linea]  += 1
        else:
            d["total_no"]          += 1
            d["lineas_no"][linea]  += 1

    # Top 20 tipos (global y por potencial)
    top_tipos    = sorted([{"tipo": t, "n": n} for t, n in por_tipo.items()],    key=lambda x: -x["n"])[:20]
    top_tipos_si = sorted([{"tipo": t, "n": n} for t, n in por_tipo_si.items()], key=lambda x: -x["n"])[:20]
    top_tipos_no = sorted([{"tipo": t, "n": n} for t, n in por_tipo_no.items()], key=lambda x: -x["n"])[:20]

    # Construir lista de clientes
    clientes = []
    for nombre, d in cli_map.items():
        ls = d["lineas"]
        # Estado más frecuente
        if d["estados"]:
            estado_cli = max(d["estados"], key=lambda e: d["estados"][e])
        else:
            estado_cli = "Sin clasificar"

        # Región del cliente: la de la mayoría de sus equipos. Sólo un cliente
        # de 1.501 (CESFAM Colbún) tiene equipos repartidos en dos regiones,
        # así que asignar la dominante no distorsiona el resumen regional.
        region_cli = (max(d["regiones"], key=lambda r: d["regiones"][r])
                      if d["regiones"] else "Sin región")

        mmq_reas = ls.get("MMQ", 0) + ls.get("REAS", 0)
        lineas_conocidas = {"DENTAL", "ESTERILIZACIÓN", "ESTERILIZACION", "INCARDIA", "ENDOSCOPIA", "MOBILIARIO CLINICO", "MMQ", "REAS"}
        otros = sum(v for k, v in ls.items() if k not in lineas_conocidas and k not in ("MMQ", "REAS"))

        ls_si  = d["lineas_si"]
        ls_no  = d["lineas_no"]
        mmq_reas_si = ls_si.get("MMQ", 0) + ls_si.get("REAS", 0)
        mmq_reas_no = ls_no.get("MMQ", 0) + ls_no.get("REAS", 0)
        otros_si = sum(v for k, v in ls_si.items() if k not in lineas_conocidas and k not in ("MMQ", "REAS"))
        otros_no = sum(v for k, v in ls_no.items() if k not in lineas_conocidas and k not in ("MMQ", "REAS"))

        clientes.append({
            "nombre":            nombre,
            "total":             d["total"],
            "total_si":          d["total_si"],
            "total_no":          d["total_no"],
            "dental":            ls.get("DENTAL", 0),
            "dental_si":         ls_si.get("DENTAL", 0),
            "dental_no":         ls_no.get("DENTAL", 0),
            "esterilizacion":    ls.get("ESTERILIZACIÓN", ls.get("ESTERILIZACION", 0)),
            "esterilizacion_si": ls_si.get("ESTERILIZACIÓN", ls_si.get("ESTERILIZACION", 0)),
            "esterilizacion_no": ls_no.get("ESTERILIZACIÓN", ls_no.get("ESTERILIZACION", 0)),
            "incardia":          ls.get("INCARDIA", 0),
            "incardia_si":       ls_si.get("INCARDIA", 0),
            "incardia_no":       ls_no.get("INCARDIA", 0),
            "endoscopia":        ls.get("ENDOSCOPÍA", ls.get("ENDOSCOPIA", 0)),
            "endoscopia_si":     ls_si.get("ENDOSCOPÍA", ls_si.get("ENDOSCOPIA", 0)),
            "endoscopia_no":     ls_no.get("ENDOSCOPÍA", ls_no.get("ENDOSCOPIA", 0)),
            "mobiliario":        ls.get("MOBILIARIO CLINICO", 0),
            "mobiliario_si":     ls_si.get("MOBILIARIO CLINICO", 0),
            "mobiliario_no":     ls_no.get("MOBILIARIO CLINICO", 0),
            "mmq_reas":          mmq_reas,
            "mmq_reas_si":       mmq_reas_si,
            "mmq_reas_no":       mmq_reas_no,
            "otros":             otros,
            "otros_si":          otros_si,
            "otros_no":          otros_no,
            "region":            region_cli,
            "estado":            estado_cli,
            "con_contrato":      estado_cli in ("Contrato", "Garantia"),
            "potencial_st":      d.get("_potencial_st", False),
        })

    # Excluir clientes internos (GEMCO)
    clientes = [c for c in clientes if "GEMCO" not in c["nombre"]]

    # Ordenar clientes por total desc
    clientes.sort(key=lambda x: -x["total"])

    # Recalcular total sin GEMCO (para que KPIs sean consistentes)
    total_sin_gemco    = sum(c["total"]    for c in clientes)
    total_si_sin_gemco = sum(c["total_si"] for c in clientes)
    total_no_sin_gemco = sum(c["total_no"] for c in clientes)

    print(f"       Base Instalada: {total_sin_gemco} activos (sin GEMCO) | {len(clientes)} clientes | {len(por_tipo)} tipos")
    _sr = reg_map.get("Sin región", {}).get("total", 0)
    print(f"       BI por region: {len(reg_map)} regiones | {_sr} equipos sin region")
    por_tipo_linea_out = {
        linea: sorted([{"tipo": t, "n": n} for t, n in ctr.items()], key=lambda x: -x["n"])[:6]
        for linea, ctr in tipo_por_linea.items()
    }

    return {
        "total":          total_sin_gemco,
        "total_si":       total_si_sin_gemco,
        "total_no":       total_no_sin_gemco,
        "por_linea":      dict(por_linea),
        "por_tipo":       top_tipos,
        "por_tipo_si":    top_tipos_si,
        "por_tipo_no":    top_tipos_no,
        "por_tipo_linea": por_tipo_linea_out,
        "por_estado":     dict(por_estado),
        "clientes":       clientes,
        "por_region": {
            r: {"total": d["total"], "total_si": d["total_si"],
                "n_clientes": len(d["_clientes"]),
                "lineas": dict(d["lineas"]), "lineas_si": dict(d["lineas_si"])}
            for r, d in sorted(reg_map.items(), key=lambda x: -x[1]["total"])
        },
    }


# Nombres de región de BASE INSTALADA normalizados a los que usa el resto del
# panel (BASE MAPA), para que el color y la posición coincidan entre hojas.
_BI_REGION_ALIAS = {
    "LA ARAUCANIA": "Araucanía", "ARAUCANIA": "Araucanía",
    "BIOBIO": "Bío Bío", "BIO BIO": "Bío Bío", "BIO-BIO": "Bío Bío",
    "METROPOLITANA": "Metropolitana", "REGION METROPOLITANA": "Metropolitana",
    "O'HIGGINS": "O'Higgins", "OHIGGINS": "O'Higgins",
    "MAGALLANES": "Magallanes", "NUBLE": "Ñuble",
    "SIN INFORMACION": "Sin región", "0": "Sin región", "#N/A": "Sin región",
}


def _norm_cli(s):
    """Normaliza nombres de cliente para cruzarlos entre hojas:
    mayúsculas, espacios colapsados y sin tildes."""
    if not s:
        return ""
    t = " ".join(str(s).strip().upper().split())
    return "".join(c for c in unicodedata.normalize("NFD", t)
                   if unicodedata.category(c) != "Mn")


def _norm_region(s):
    """Deja el nombre de región tal como lo usa el resto del panel."""
    t = safe_str(s).strip()
    if not t:
        return "Sin región"
    k = "".join(c for c in unicodedata.normalize("NFD", t.upper())
                if unicodedata.category(c) != "Mn")
    if k in _BI_REGION_ALIAS:
        return _BI_REGION_ALIAS[k]
    return t


# ══════════════════════════════════════════════════════════════════════════════
# 5c. HOJA: ANALISIS FACTURACIÓN + GD-PPTO
#     Tabla 1 (filas 4-7): resumen mensual por linea de negocio
#     Tabla 2 (filas 11-14): desglose semanal del mes en curso
#     GD-PPTO filas 20-22: presupuesto mensual por catálogo
#
#     Valores en la hoja están en MM$ → se multiplican por 1e6 para pesos
# ══════════════════════════════════════════════════════════════════════════════
def read_analisis_fac(wb):
    # Encontrar hoja por nombre (con/sin acento)
    ws_anal = None
    ws_gd   = None
    for name in wb.sheetnames:
        nl = name.lower()
        if "analisis" in nl or "análisis" in nl:
            ws_anal = wb[name]
        if "gd" in nl and "ppto" in nl:
            ws_gd = wb[name]

    MM = 1_000_000  # los valores en la hoja están en MM$

    # ── Analisis Facturación ────────────────────────────────────────────────
    result = {
        "semana": 0, "mes_nombre": "",
        "tabla_mensual": [], "tabla_semanal": [],
        "ts_ingresos": 0.0, "ts_total_ytd": 0.0,
        "ts_ppto_acum": 0.0, "ts_ppto_anual": 0.0,
        "ts_ingresos_aa": 0.0,
        "ppto_mensual": [0.0] * 12,
    }

    if ws_anal:
        rows = list(ws_anal.iter_rows(min_row=1, max_row=35, values_only=True))

        # Fila 2 (idx 1): semana y mes
        r2 = rows[1] if len(rows) > 1 else []
        result["semana"]     = to_int(r2[4]) if len(r2) > 4 else 0
        result["mes_nombre"] = safe_str(r2[6]) if len(r2) > 6 else ""

        def _parse_mensual(row):
            if not row or not row[1]:
                return None
            return {
                "linea":            safe_str(row[1]),
                "ingresos":         to_float(row[4]) * MM,
                "provision":        to_float(row[5]) * MM,
                "gd":               to_float(row[6]) * MM,
                "total_ytd":        to_float(row[7]) * MM,
                "ppto_acum":        to_float(row[8]) * MM,
                "delta_ppto":       to_float(row[9]) * MM,
                "delta_ppto_pct":   to_float(row[10]),
                "ingresos_aa":      to_float(row[11]) * MM,
                "delta_aa":         to_float(row[12]) * MM,
                "delta_aa_pct":     to_float(row[13]),
                "ppto_anual":       to_float(row[14]) * MM,
            }

        def _parse_semanal(row):
            if not row or not row[1]:
                return None
            return {
                "linea":        safe_str(row[1]),
                "ingresos":     to_float(row[4]) * MM,
                "s1":           to_float(row[5]) * MM,
                "s2":           to_float(row[6]) * MM,
                "s3":           to_float(row[7]) * MM,
                "s4":           to_float(row[8]) * MM,
                "s5":           to_float(row[9]) * MM,
                "acum_mes":     to_float(row[10]) * MM,
                "ppto_mes":     to_float(row[11]) * MM,
                "delta_ppto":   to_float(row[12]) * MM,
                "delta_ppto_pct": to_float(row[13]),
                "vs_mes_ant":   to_float(row[14]) * MM if len(row) > 14 else 0,
                "delta_vs_ant": to_float(row[15]) * MM if len(row) > 15 else 0,
            }

        # Tabla 1: filas 4-7 (índice 3-6)
        for idx in [3, 4, 5, 6]:
            parsed = _parse_mensual(rows[idx] if idx < len(rows) else None)
            if parsed:
                result["tabla_mensual"].append(parsed)

        # Tabla 2: filas 11-14 (índice 10-13)
        for idx in [10, 11, 12, 13]:
            parsed = _parse_semanal(rows[idx] if idx < len(rows) else None)
            if parsed:
                result["tabla_semanal"].append(parsed)

        # Totales TS (fila de "Ingresos Totales")
        for r in result["tabla_mensual"]:
            if "total" in r["linea"].lower():
                result["ts_ingresos"]    = r["ingresos"]
                result["ts_total_ytd"]   = r["total_ytd"]
                result["ts_ppto_acum"]   = r["ppto_acum"]
                result["ts_ppto_anual"]  = r["ppto_anual"]
                result["ts_ingresos_aa"] = r["ingresos_aa"]
                break

    # ── Líneas de negocio mensuales: filas 31-34 (índices 30-33 en rows) ────
    result["lineas_mensual"] = {}
    if ws_anal:
        for row in rows[30:35]:
            label = safe_str(row[1]).strip() if row[1] else ""
            if not label:
                continue
            meses = []
            for m in range(12):
                v = row[2 + m] if len(row) > 2 + m else 0
                meses.append(round(to_float(v)))
            result["lineas_mensual"][label] = meses

    # ── GD-PPTO: presupuesto mensual (fila 23 = total TECSERVICE) ───────────
    if ws_gd:
        # Fila 23 = total TECSERVICE (suma de REAS + ST + Trazabilidad)
        gd_rows = list(ws_gd.iter_rows(min_row=23, max_row=23, values_only=True))
        ppto_mes = [0.0] * 12
        for row in gd_rows:
            if not row:
                continue
            for m in range(12):
                col_idx = 5 + m  # Enero = col F (índice 5)
                if col_idx < len(row):
                    ppto_mes[m] = to_float(row[col_idx]) * MM
        result["ppto_mensual"] = [round(v) for v in ppto_mes]

    return result


# ══════════════════════════════════════════════════════════════════════════════
# 6. HOJA: BASE MAPA (hoja "BASE MAPA" del Excel principal)
#    Columnas (0-based):
#    0=NombreCliente  1=TipoCliente  2=Ingresos2025  3=Ingresos2026  4=BiTotal
#    6=EqEster  7=EqDental  8=EqEndo  9=EqIncardia  10=EqMMQ  11=EqREAS
#    12=EqMobClin  13=EqMobOtros
#    14=Región  15=Comuna
#    17=PipeEster  18=PipeEndo  19=PipeDental
#    20=PotencialEquipos  21=PotSTGarantía  22=PotContratosBIActual
#    23=PotSTTotal  24=Latitud  25=Longitud  26=Contrato(Sí/No)
#    contratos y sat se enriquecen en enrich_mapa_data(); cc viene directo de col AA
# ══════════════════════════════════════════════════════════════════════════════
def read_mapa(wb):
    ws = None
    for name in wb.sheetnames:
        if name.strip().upper() == "BASE MAPA":
            ws = wb[name]
            break
    if ws is None:
        print("  ADVERTENCIA: no se encontró la hoja 'BASE MAPA' en el Excel.")
        return []
    clientes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre = safe_str(row[0])
        if not nombre:
            continue
        # Latitud/longitud pueden venir vacías: el cliente igual entra al
        # panel — con su ingreso, región y potencial — y sólo se queda fuera
        # de las burbujas del mapa, que son las únicas que necesitan el punto.
        lat = to_float(row[24], None)   # col Y
        lon = to_float(row[25], None)   # col Z
        if lat == 0 or lon == 0:
            lat = lon = None
        region_raw = safe_str(row[14])
        region = region_raw.replace("Región ", "").replace("Region ", "").strip()

        # Potencial Equipos (col U) = sum pipeline por línea (R+S+T)
        pot_eq_ester  = to_int(row[17])   # col R: Pipeline Esterilización
        pot_eq_endo   = to_int(row[18])   # col S: Pipeline Endoscopía
        pot_eq_dental = to_int(row[19])   # col T: Pipeline Dental
        pot_eq        = to_int(row[20])   # col U: Potencial Equipos total

        # Potencial ST (col V + col W; col X = total precalculado)
        pot_st_gar   = to_int(row[21])   # col V: Potencial ST Garantía
        pot_st_contr = to_int(row[22])   # col W: Potencial Contratos sobre BI Actual
        pot_st       = to_int(row[23])   # col X: Potencial ST Total (directo del Excel)
        if pot_st == 0:
            pot_st = pot_st_gar + pot_st_contr  # fallback si col X está vacía

        # Contrato (col AA = index 26): fuente autorativa desde Excel
        cc_raw = safe_str(row[26]).strip().upper() if len(row) > 26 and row[26] is not None else ""
        cc = cc_raw in ("SI", "SÍ", "S", "1", "TRUE", "VERDADERO", "X")

        # Satisfacción (col AB = index 27): escala 1–5, 0 = sin dato
        sat_raw = row[27] if len(row) > 27 else None
        sat = None
        if sat_raw is not None:
            try:
                sat_val = int(float(sat_raw))
                sat = sat_val if 1 <= sat_val <= 5 else None
            except Exception:
                sat = None

        eq = {
            "Esterilización": to_int(row[6]),
            "Dental":         to_int(row[7]),
            "Endoscopía":     to_int(row[8]),
            "Incardia":       to_int(row[9]),
            "MMQ":            to_int(row[10]),
            "REAS":           to_int(row[11]),
            "Mob.Clínico":    to_int(row[12]),
            "Mob.Otros":      to_int(row[13]),
        }
        clientes.append({
            "n":            nombre,
            "tipo":         safe_str(row[1]),
            "ingreso_2025": to_int(row[2]),    # col C
            "ingreso_2026": to_int(row[3]),    # col D
            "ingreso":      to_int(row[3]),    # alias 2026
            "bi":           to_int(row[4]),    # BI Total (equipos con pot. ST)
            "eq":           eq,
            "region":       region,
            "comuna":       safe_str(row[15]),
            "contratos":    0,                 # enrich_mapa_data
            "pipe":         pot_eq,
            "lat":          round(lat, 7) if lat is not None else None,
            "lon":          round(lon, 7) if lon is not None else None,
            "cc":           cc,                # desde col AA del Excel
            "margen":       0,
            "sat":          sat,               # desde col AB del Excel
            "pot_eq":       pot_eq,
            "pot_eq_ester": pot_eq_ester,
            "pot_eq_endo":  pot_eq_endo,
            "pot_eq_dental": pot_eq_dental,
            "pot_st":       pot_st,
            "pot_st_gar":   pot_st_gar,
            "pot_st_contr": pot_st_contr,
            "pot":          pot_eq + pot_st,
        })
    return clientes


# ══════════════════════════════════════════════════════════════════════════════
# 6b. HOJA: CASOS RELEVANTES
#     Dos tablas lado a lado:
#     Tabla izquierda A-F (cols 0-5): Casos Relevantes
#     Tabla derecha   J-T (cols 9-19): Equipos Detenidos
# ══════════════════════════════════════════════════════════════════════════════
def read_casos(wb):
    """Casos relevantes y equipos detenidos.

    La hoja tiene dos tablas lado a lado y ya se le han insertado columnas
    tres veces (ESTADO y Acciones a la izquierda; Fecha Ingreso Caso, Fecha
    de ingreso y Costo CIF a la derecha), corriendo todo lo que venía después
    y dejando el panel leyendo campos equivocados en silencio. Por eso las
    columnas se ubican por su encabezado y no por un índice fijo.
    """
    ws = None
    for name in wb.sheetnames:
        if "caso" in name.lower() and "relevante" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        return {"casos": [], "equipos": []}

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return {"casos": [], "equipos": []}
    cab = filas[0]

    def buscar(*fragmentos):
        """Índice de la primera columna cuyo encabezado contenga el fragmento."""
        for frag in fragmentos:
            f = _norm_cli(frag)
            for i, v in enumerate(cab):
                if v is None:
                    continue
                if f in _norm_cli(v):
                    return i
        return None

    # Tabla izquierda: casos
    C = {
        "coordinador": buscar("COORDINADOR"),
        "cliente":     buscar("CLIENTE"),
        "problema":    buscar("PROBLEMA"),
        "estado":      buscar("ESTADO"),
        "responsable": buscar("RESPONSABLE"),
        "comentario":  buscar("COMENTARIO"),
        "salesforce":  buscar("REGISTRO EN SALESFORCE", "SALESFORCE"),
        "acciones":    buscar("ACCIONES"),
        "fecha_caso":  buscar("FECHA INGRESO CASO"),
    }
    # Tabla derecha: equipos. Se busca desde la columna del modelo hacia la
    # derecha para no confundir "Estado" del caso con "Estado" del equipo, ni
    # "Cliente" con "Nombre Cliente".
    i_mod = buscar("MODELO")
    def buscarD(*fragmentos):
        for frag in fragmentos:
            f = _norm_cli(frag)
            for i, v in enumerate(cab):
                if v is None or (i_mod is not None and i < i_mod):
                    continue
                if f in _norm_cli(v):
                    return i
        return None
    E = {
        "modelo":           i_mod,
        "nombre":           buscarD("NOMBRE DE ACTIVO"),
        "serie":            buscarD("NUMERO DE SERIE"),
        "marca":            buscarD("MARCA"),
        "estado":           buscarD("ESTADO"),
        "coordinadora":     buscarD("COORDINADORA"),
        "comentario_coord": buscarD("COMENTARIO COORDINADORA"),
        "comentario_cat":   buscarD("COMENTARIO COORDINADORA 2"),
        "comentario_mat":   buscarD("COMENTARIO MATILDE"),
        "contrato_num":     buscarD("NUMERO DE CONTRATO"),
        "garantia":         buscarD("ESTADO GARANTIA"),
        "nombre_cliente":   buscarD("NOMBRE CLIENTE"),
        "neta_mes":         buscarD("FACTURACION NETA MES"),
        "fac_anual":        buscarD("FACTURACION ANUAL"),
        "fac_ytd":          buscarD("FACTURACION A LA FECHA"),
        "fecha_inicio":     buscarD("FECHA INC"),   # "Fecha Inciio" / "Fecha Incio Contrato"
        "fecha_fin":        buscarD("FECHA FIN"),
        "fecha_ingreso":    buscarD("FECHA DE INGRESO"),
        "costo_cif":        buscarD("COSTO CIF"),
        # «Fecha de ingreso» es la del equipo; «Fecha de Estado» es la del
        # estado actual. Son cosas distintas: la primera mide la vida del
        # equipo, la segunda cuanto lleva detenido en la etapa en que esta.
        "fecha_estado":     buscarD("FECHA DE ESTADO"),
        "dias_estado":      buscarD("ANTIGUEDAD DE ESTADO"),
    }
    # El estado del equipo se toma de «COMENTARIO MATILDE», que es la columna
    # categorizada que se muestra y por la que se filtra. Las otras dos
    # columnas de comentario quedan como texto libre de apoyo. Si algún día
    # falta esa columna, se cae a la que tenga menos valores distintos, que
    # es la forma que tiene la categorizada.
    if E["comentario_mat"] is not None:
        E["comentario_cat"], E["comentario_mat"] = E["comentario_mat"], E["comentario_cat"]
    else:
        cols_com = sorted({i for i in (E["comentario_coord"], E["comentario_cat"]) if i is not None})
        if cols_com:
            distintos = {i: len({_norm_cli(r[i]) for r in filas[1:]
                                 if r and i < len(r) and r[i] not in (None, "")})
                         for i in cols_com}
            i_cat = min(distintos, key=lambda i: distintos[i])
            E["comentario_cat"] = i_cat
            libres = [i for i in cols_com if i != i_cat]
            E["comentario_coord"] = libres[0] if libres else None

    faltan = [k for k, v in E.items() if v is None and
              k not in ("fecha_ingreso", "costo_cif", "fecha_estado", "dias_estado")]
    if faltan:
        print(f"  ADVERTENCIA: columnas no encontradas en Casos Relevantes: {', '.join(faltan)}")

    def val(row, i):
        return row[i] if (i is not None and i < len(row)) else None
    def txt(row, i):
        v = val(row, i)
        return safe_str(v) if v is not None else ""
    def num(row, i):
        v = val(row, i)
        return to_float(v) if isinstance(v, (int, float)) else 0
    def fecha(row, i):
        v = val(row, i)
        return v.strftime("%d-%m-%Y") if isinstance(v, (datetime, date)) else ""
    def sin_asoc(row, i):
        t = txt(row, i)
        return "" if "NO ASOCIADO" in t.upper() else t

    casos, equipos = [], []
    last_coord = ""
    # Las dos tablas están lado a lado en la MISMA fila del Excel, así que el
    # número de fila las une: el caso trae el cliente abreviado a mano («HP
    # Tisné») y el equipo de esa misma fila trae el nombre completo, que es el
    # que permite cruzar contra facturación y base instalada.
    for _i, row in enumerate(filas[1:]):
        if not row:
            continue
        # El coordinador puede venir en celdas combinadas: se arrastra
        c0 = txt(row, C["coordinador"])
        if c0:
            last_coord = c0

        cliente = txt(row, C["cliente"])
        if cliente:
            casos.append({
                "coordinador":  last_coord,
                "cliente":      cliente,
                "problema":     txt(row, C["problema"]),
                "estado":       txt(row, C["estado"]),
                "responsable":  txt(row, C["responsable"]),
                "comentario":   txt(row, C["comentario"]),
                "salesforce":   txt(row, C["salesforce"]),
                "acciones":     txt(row, C["acciones"]),
                "fecha_caso":   fecha(row, C["fecha_caso"]),
                "fila":         _i,
            })

        modelo = txt(row, E["modelo"])
        if modelo:
            equipos.append({
                "modelo":           modelo,
                "nombre":           txt(row, E["nombre"]),
                "serie":            txt(row, E["serie"]),
                "marca":            txt(row, E["marca"]),
                "estado":           txt(row, E["estado"]),
                "coordinadora":     txt(row, E["coordinadora"]),
                "comentario_coord": txt(row, E["comentario_coord"]),
                "comentario_cat":   txt(row, E["comentario_cat"]),
                "comentario_mat":   txt(row, E["comentario_mat"]),
                "contrato_num":     txt(row, E["contrato_num"]),
                "garantia":         txt(row, E["garantia"]),
                "nombre_cliente":   sin_asoc(row, E["nombre_cliente"]),
                "neta_mes":         num(row, E["neta_mes"]),
                "fac_anual":        num(row, E["fac_anual"]),
                "fac_ytd":          num(row, E["fac_ytd"]),
                "fecha_inicio":     fecha(row, E["fecha_inicio"]),
                "fecha_fin":        fecha(row, E["fecha_fin"]),
                "fecha_ingreso":    fecha(row, E["fecha_ingreso"]),
                "costo_cif":        num(row, E["costo_cif"]),
                "fila":             _i,
                "fecha_estado":     fecha(row, E["fecha_estado"]),
                # La antiguedad del Excel es un =HOY()-fecha, asi que llega
                # cuadrada al dia de la conversion; el panel igual la
                # recalcula desde la fecha para no depender de eso.
                "dias_estado":      num(row, E["dias_estado"]),
            })

    return {"casos": casos, "equipos": equipos}



def _ffill(valor, previo):
    """Los export de Salesforce dejan en blanco las filas que continúan el
    mismo grupo (propietario, línea, orden). Arrastra el último valor visto."""
    v = safe_str(valor).strip()
    return v if v else previo


def _parse_fecha(v):
    """Fecha que puede venir como datetime o como texto D/M/YYYY."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = safe_str(v).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def read_brecha_oport(wb):
    """Oportunidades adjudicadas que aún no se facturan.

    Fuente: hoja "Brecha Oport por Facturar". El monto es "Monto adjudicado
    (convertido)" (col H). El propietario (col A) viene con celdas en blanco
    en las filas que continúan el mismo grupo, así que se arrastra.
    """
    ws = None
    for name in wb.sheetnames:
        n = name.strip().lower()
        if "brecha" in n and "oport" in n:
            ws = wb[name]
            break
    if ws is None:
        return {}

    items, prop = [], ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 9:
            continue
        prop = _ffill(row[0], prop)
        cliente = safe_str(row[2]).strip()
        monto   = to_float(row[7])
        if not cliente and monto == 0:
            continue
        f = _parse_fecha(row[6])
        items.append({
            "prop":     prop or "Sin asignar",
            "cliente":  cliente or "(sin cliente)",
            "oport":    safe_str(row[3]).strip(),
            "contrato": safe_str(row[4]).strip(),
            "oc":       safe_str(row[5]).strip(),
            "fecha":    f.isoformat() if f else "",
            "fecha_fmt": f.strftime("%d/%m/%Y") if f else "",
            "ov":       safe_str(row[8]).strip(),
            "monto":    round(monto),
        })

    if not items:
        return {}
    items.sort(key=lambda x: -x["monto"])

    def agrupa(campo):
        g = defaultdict(lambda: [0.0, 0])
        for it in items:
            g[it[campo]][0] += it["monto"]
            g[it[campo]][1] += 1
        return [{"k": k, "monto": round(v[0]), "n": v[1]}
                for k, v in sorted(g.items(), key=lambda x: -x[1][0])]

    return {
        "total":          round(sum(i["monto"] for i in items)),
        "n":              len(items),
        "por_propietario": agrupa("prop"),
        "por_cliente":     agrupa("cliente"),
        "items":           items,
    }


# Etiquetas de rotación canónicas. La columna P de "Brecha Sin Stock" mezcla
# mayúsculas y tildes ("Sin Información" / "Sin información"), y sin unificar
# el resumen abre dos categorías para lo mismo.
_ROT_CANON = {
    "ALTA ROTACION":    "Alta Rotacion",
    "MEDIANA ROTACION": "Mediana Rotacion",
    "BAJA ROTACION":    "Baja Rotacion",
    "SIN ROTACION":     "Sin Rotacion",
    "SIN INFORMACION":  "Sin Información",
}


def _rot_canon(valor):
    txt = safe_str(valor).strip() if valor is not None else ""
    if not txt:
        return "Sin Información"
    return _ROT_CANON.get(_norm_cli(txt), txt)


def read_brecha_stock(wb):
    """Oportunidades detenidas por falta de stock de repuestos.

    Fuente: hoja "Brecha Sin Stock". El monto es "Precio total" (col N).
    Propietario (A), línea de negocio (B) y orden de venta (C) vienen con
    celdas en blanco en las filas de continuación y se arrastran.
    """
    ws = None
    for name in wb.sheetnames:
        n = name.strip().lower()
        if "brecha" in n and "stock" in n:
            ws = wb[name]
            break
    if ws is None:
        return {}

    items = []
    prop = linea = ov = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 14:
            continue
        prop  = _ffill(row[0], prop)
        linea = _ffill(row[1], linea)
        ov    = _ffill(row[2], ov)
        monto = to_float(row[13])
        cod   = safe_str(row[9]).strip()
        # La hoja cierra con una fila "Total" que repite la suma de la columna
        # N; sin descartarla la brecha salía al doble. Toda línea real trae
        # código de producto — las de continuación heredan propietario y OV,
        # pero nunca el código.
        if not cod:
            continue
        if safe_str(row[0]).strip().upper() in ("TOTAL", "TOTALES"):
            continue
        f = _parse_fecha(row[8])
        items.append({
            "prop":     prop or "Sin asignar",
            "linea":    linea or "Sin línea",
            "ov":       ov,
            "oport":    safe_str(row[4]).strip(),
            "cliente":  safe_str(row[6]).strip() or "(sin cliente)",
            "oc":       safe_str(row[7]).strip(),
            "fecha":    f.isoformat() if f else "",
            "fecha_fmt": f.strftime("%d/%m/%Y") if f else "",
            "cod":      cod,
            "prod":     safe_str(row[10]).strip(),
            "cant":     round(to_float(row[11]), 2),
            "pu":       round(to_float(row[12])),
            "monto":    round(monto),
            "rot":      _rot_canon(row[15] if len(row) > 15 else None),
            "dias":     (TODAY - f).days if f else None,
        })

    if not items:
        return {}
    items.sort(key=lambda x: -x["monto"])

    def agrupa(campo):
        g = defaultdict(lambda: [0.0, 0, 0.0])
        for it in items:
            g[it[campo]][0] += it["monto"]
            g[it[campo]][1] += 1
            g[it[campo]][2] += it["cant"]
        return [{"k": k, "monto": round(v[0]), "n": v[1], "cant": round(v[2], 2)}
                for k, v in sorted(g.items(), key=lambda x: -x[1][0])]

    # Antigüedad: cuánto llevan esperando repuesto. Es lo más accionable —
    # una oportunidad de hace 8 meses probablemente ya se perdió.
    TRAMOS = [(0, 30, "0–30 días"), (31, 60, "31–60 días"), (61, 90, "61–90 días"),
              (91, 180, "91–180 días"), (181, 10**6, "Más de 180 días")]
    aging = []
    for lo, hi, lbl in TRAMOS:
        sel = [i for i in items if i["dias"] is not None and lo <= i["dias"] <= hi]
        aging.append({"k": lbl, "monto": round(sum(i["monto"] for i in sel)), "n": len(sel)})
    sin_fecha = [i for i in items if i["dias"] is None]
    if sin_fecha:
        aging.append({"k": "Sin fecha", "monto": round(sum(i["monto"] for i in sin_fecha)),
                      "n": len(sin_fecha)})

    # Mes de creación de la oportunidad
    MESES_ABR = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
                 "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    g_mes = defaultdict(lambda: [0.0, 0])
    for it in items:
        if not it["fecha"]:
            continue
        y, m = it["fecha"][:4], int(it["fecha"][5:7])
        g_mes[(y, m)][0] += it["monto"]
        g_mes[(y, m)][1] += 1
    por_mes = [{"k": f"{MESES_ABR[m-1]} {y[2:]}", "monto": round(v[0]), "n": v[1]}
               for (y, m), v in sorted(g_mes.items(), key=lambda x: (x[0][0], x[0][1]))]

    # Productos que más veces bloquean una venta
    g_prod = defaultdict(lambda: [0.0, 0, 0.0, "", set()])
    for it in items:
        e = g_prod[it["cod"]]
        e[0] += it["monto"]; e[1] += 1; e[2] += it["cant"]
        if not e[3]:
            e[3] = it["prod"]
        e[4].add(it["cliente"])
    # Back order: unidades ya pedidas al proveedor para ese SKU.
    # Hoja "Back Order", col G = SKU, col K = Cantidad Solicitada. Se suman
    # todas las filas sin mirar "Estatus OC": ese campo indica si la orden
    # está pagada, no si la mercadería llegó, así que no hay doble conteo
    # contra el stock de bodega.
    bo_sol = defaultdict(float)
    ws_bo = next((wb[n] for n in wb.sheetnames if n.strip().lower() == "back order"), None)
    if ws_bo is not None:
        for row in ws_bo.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 11:
                continue
            sku = safe_str(row[6]).strip()
            if sku:
                bo_sol[sku] += to_float(row[10])

    productos = [{"cod": k, "prod": v[3], "monto": round(v[0]), "n": v[1],
                  "cant": round(v[2], 2), "n_cli": len(v[4]),
                  "bo": round(bo_sol.get(k, 0.0), 2)}
                 for k, v in sorted(g_prod.items(), key=lambda x: -x[1][0])]

    # Rotación del SKU (col P). Se usa para el resumen general y para el
    # desglose por cliente de la tabla expandible.
    rot_g = defaultdict(lambda: [0.0, 0])
    for it in items:
        rot_g[it["rot"]][0] += it["monto"]
        rot_g[it["rot"]][1] += 1
    _tm = sum(v[0] for v in rot_g.values()) or 1
    _tn = sum(v[1] for v in rot_g.values()) or 1
    por_rotacion = [
        {"k": k, "monto": round(v[0]), "n": v[1],
         "pct_monto": round(v[0] / _tm * 100, 1), "pct_n": round(v[1] / _tn * 100, 1)}
        for k, v in sorted(rot_g.items(), key=lambda x: -x[1][0])
    ]

    # Clientes con su detalle de SKU y su mezcla de rotación
    cli_det = defaultdict(lambda: {"monto": 0.0, "n": 0, "cant": 0.0,
                                   "rot": defaultdict(lambda: [0.0, 0]), "skus": []})
    for it in items:
        d = cli_det[it["cliente"]]
        d["monto"] += it["monto"]
        d["n"]     += 1
        d["cant"]  += it["cant"]
        d["rot"][it["rot"]][0] += it["monto"]
        d["rot"][it["rot"]][1] += 1
        d["skus"].append({k: it[k] for k in
                          ("cod", "prod", "cant", "pu", "monto", "rot", "dias", "fecha_fmt", "linea", "oport")})
    clientes_det = []
    for c, d in sorted(cli_det.items(), key=lambda x: -x[1]["monto"]):
        d["skus"].sort(key=lambda x: -x["monto"])
        clientes_det.append({
            "cliente": c,
            "monto":   round(d["monto"]),
            "n":       d["n"],
            "cant":    round(d["cant"], 2),
            "rot":     {k: {"monto": round(v[0]), "n": v[1],
                            "pct": round(v[0] / d["monto"] * 100, 1) if d["monto"] else 0}
                        for k, v in sorted(d["rot"].items(), key=lambda x: -x[1][0])},
            "skus":    d["skus"],
        })

    dias_val = [i["dias"] for i in items if i["dias"] is not None]
    return {
        "por_rotacion":  por_rotacion,
        "clientes_det":  clientes_det,
        "total":           round(sum(i["monto"] for i in items)),
        "n":               len(items),
        "n_ov":            len({i["ov"] for i in items if i["ov"]}),
        "n_clientes":      len({i["cliente"] for i in items}),
        "cant_total":      round(sum(i["cant"] for i in items), 2),
        "dias_prom":       round(sum(dias_val) / len(dias_val)) if dias_val else None,
        "dias_max":        max(dias_val) if dias_val else None,
        "por_linea":       agrupa("linea"),
        "por_cliente":     agrupa("cliente"),
        "por_propietario": agrupa("prop"),
        "aging":           aging,
        "por_mes":         por_mes,
        "productos":       productos,
        "items":           items,
    }


# ══════════════════════════════════════════════════════════════════════════════
# EQUIPOS QUE MÁS FALLAN
# ══════════════════════════════════════════════════════════════════════════════
# La hoja "Repuestos Vendidas" no tiene una columna de equipo: lo que hay es el
# «Nombre de cotización» (col F), texto libre escrito a mano por quien cotizó.
# Ahí aparece el equipo cuando la cotización es de una máquina concreta
# ("AUTOCLAVE STEELCO VS 12/2 VDX 12114 OT 57902") y no aparece cuando la
# cotización es masiva ("M. PREVENTIVOS DICIEMBRE - HOSPITAL BASE DE OSORNO")
# o es una cuota de convenio. Por eso el equipo se deduce por patrones y se
# reconoce explícitamente lo que no se pudo identificar, en vez de repartirlo.
#
# Se emite a nivel de línea, no agregado: son ~3.000 filas y así el panel puede
# cruzar cualquier combinación de año, naturaleza y marca contando cotizaciones
# y clientes distintos sin que las sumas se dupliquen.

_EQ_TIPOS = [
    ("Lavadora ultrasónica",      r"ULTRASONIC|ULTRASONID|SONICA|\bUS\s?\d{2,3}\b"),
    ("Lavadora de endoscopios",   r"REPROCESAD|LAVAENDOSCOPI|\bWD440\b"),
    ("Lavadora descontaminadora", r"DESCONTAMINAD|\bDS\s?\d{3}|\bPG\s?\d{4}|LAVADORA|LAVAINSTRUMENT"),
    ("Secadora",                  r"SECADORA|SECADO"),
    ("Autoclave",                 r"AUTOCLAVE|ESTERILIZADOR|VACU[CK]LAV|CLINICLAVE|"
                                  r"\bE?VS\s?\d{1,2}\s?/|\bLVS\b|\bEVS\b"),
    ("Endoscopio",                r"ENDOSCOPI|GASTROSCOPI|COLONOSCOPI|DUODENOSCOPI|BRONCOSCOPI|"
                                  r"LARINGOSCOPI|NASOFIBRO|FIBRONASO|CISTOSCOPI|VIDEOPROCESAD|FUENTE DE LUZ"),
    ("Selladora",                 r"SELLADORA|MELASEAL"),
    ("Maceradora",                r"MACERAD|PUL[PM]MATIC|PULMATIC"),
    ("Planta de agua / ósmosis",  r"OSMOSIS|PLANTA DE AGUA|DESMINERALIZ|DESTILAD|ABLANDADOR|TRATAMIENTO DE AGUA"),
    ("Equipo dental",             r"\bDENTAL\b|TURBINA|MICROMOTOR|CHIROPRO|SILLON|UNIDAD ODONTOLOG|"
                                  r"COMPRESOR|AMALGAM|RAYOS X|RADIOGRAF"),
    ("Incubadora / cuna",         r"\bCUNA\b|INCUBADOR|FOTOTERAPIA"),
    ("Monitor / diagnóstico",     r"MONITOR|DESFIBRILAD|ELECTROBISTURI|OXIMETR|CENTRIFUG"),
    ("Lavacarros",                r"LAVACARRO|LAVA CARRO"),
]
# El modelo se busca en el mismo texto. El orden importa: las series de
# Steelco son las más frecuentes y las más específicas.
_EQ_MODELOS = [
    r"\b(?:EVS|LVS|VSX|ESX|EDX|VDX|VS)\s?\d{1,3}\s?(?:[/-]\s?\d{1,2})?\b",
    r"\bDS\s?\d{3,4}(?:\s?-\s?\d?[A-Z]{1,3}(?:-\d?[A-Z]?)?)?\b",
    r"\bPG\s?\d{4}\b",
    r"\b(?:ED|EG|EC|EPK|EPM)\s?-?\s?\d{3,4}\s?[A-Z]{0,3}\b",
    r"\bUS\s?\d{2,3}\b",
    r"\bVACU[CK]LAVE?\s?\d{2,3}\s?[A-Z]?\b",
    r"\bCLINICLAVE\s?\d{2,3}\b",
    r"\bMELASEAL\s?[A-Z]*\b",
    r"\bWD\s?\d{3}\b",
]
# Naturaleza del gasto. Sólo «Correctivo» habla derechamente de una falla; el
# resto se separa para no mezclar mantención programada con reparación.
_EQ_NATS = [
    ("Correctivo",  r"CORRECTIV|REPARACION|REPARAR|FALLA|EMERGENC|AVERIA|"
                    r"CAMBIO DE (?:REPUESTO|PIEZA)|NO ENCIENDE|FUGA"),
    ("Garantía",    r"GARANTIA"),
    ("Preventivo",  r"PREVENTIV|\bMP\b|\bM\s?\.\s?P\b|MANTENCION|MANTENIMIENTO"),
    ("Convenio",    r"CONVENIO|CUOTA|ARRIENDO"),
    ("Uso interno", r"USO INTERNO|HERRAMIENTA|BODEGA|STOCK"),
]
_EQ_FAM_V = {"AUTOCLAVES": "Autoclave", "LAVADORA": "Lavadora descontaminadora"}
_EQ_SIN = "Sin equipo identificado"


def _eq_norm(s):
    t = unicodedata.normalize("NFD", safe_str(s).upper())
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t).strip()


def read_equipos_fallas(wb):
    """Repuestos vendidos abiertos por equipo, deducido del nombre de cotización."""
    ws = None
    for name in wb.sheetnames:
        n = name.strip().lower()
        if "repuesto" in n and "vend" in n:
            ws = wb[name]
            break
    if ws is None:
        return {}

    rx_tipo = [(t, re.compile(p)) for t, p in _EQ_TIPOS]
    rx_mod = [re.compile(p) for p in _EQ_MODELOS]
    rx_nat = [(t, re.compile(p)) for t, p in _EQ_NATS]

    marcas, tipos, modelos, nats, cots, clis = [], [], [], [], [], []
    idx_m, idx_t, idx_mo, idx_n, idx_c, idx_cl = {}, {}, {}, {}, {}, {}

    def ref(lista, indice, valor):
        if valor not in indice:
            indice[valor] = len(lista)
            lista.append(valor)
        return indice[valor]

    filas_out = []
    anios = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 22:
            continue
        marca = safe_str(row[19]).strip().upper()
        if not marca:
            continue
        monto = to_float(row[15]) if isinstance(row[15], (int, float)) else 0.0
        cant = to_float(row[12]) if isinstance(row[12], (int, float)) else 0.0
        anio = row[17] if isinstance(row[17], (int, float)) else None
        mes = row[18] if isinstance(row[18], (int, float)) else None
        if anio is None or mes is None:
            continue
        anio, mes = int(anio), int(mes)
        anios.add(anio)

        txt = _eq_norm(row[5])
        tipo = None
        for nombre, rx in rx_tipo:
            if rx.search(txt):
                tipo = nombre
                break
        if tipo is None:
            # La cotización no nombra el equipo. «Equipo Asociado» (col V) se
            # deduce del producto, así que sirve de respaldo grueso.
            tipo = _EQ_FAM_V.get(safe_str(row[21]).strip().upper())
        modelo = ""
        for rx in rx_mod:
            m = rx.search(txt)
            if m:
                modelo = re.sub(r"\s*([/-])\s*", r"\1", m.group(0)).strip()
                break
        nat = "Sin clasificar"
        for nombre, rx in rx_nat:
            if rx.search(txt):
                nat = nombre
                break

        filas_out.append([
            ref(marcas, idx_m, marca),
            ref(tipos, idx_t, tipo or _EQ_SIN),
            ref(modelos, idx_mo, modelo),
            ref(nats, idx_n, nat),
            anio, mes,
            ref(cots, idx_c, safe_str(row[6]).strip() or ("#" + str(len(filas_out)))),
            ref(clis, idx_cl, safe_str(row[2]).strip()),
            round(monto),
            round(cant, 2),
        ])

    return {
        "marcas":  marcas,
        "tipos":   tipos,
        "modelos": modelos,
        "nats":    nats,
        "clientes": clis,
        "anios":   sorted(anios),
        "sin_eq":  _EQ_SIN,
        # [marca, tipo, modelo, naturaleza, año, mes, cotización, cliente, monto, cantidad]
        "filas":   filas_out,
    }


def read_repuestos_vendidos(wb):
    """Repuestos vendidos por marca, mes y cliente.

    Fuente: hoja "Repuestos Vendidas" (una fila por línea de cotización).
    Replica la tabla dinámica "TD": suma "Precio de venta" agrupando por
    "Marca 2" (marca normalizada) y año/mes. NO filtra por Estado — la
    dinámica incluye Aprobado, En borrador y Rechazado (validado: el total
    da 1.240.069.989,55 sólo sin filtrar).

    La marca se pasa a mayúsculas antes de agrupar porque el Excel trae
    PURYTAS/Purytas y NACIONAL/Nacional como variantes de la misma marca;
    la dinámica las junta y sin normalizar los totales no cuadran.

    Columnas: C=Nombre del cliente, M=Cantidad, P=Precio de venta,
              R=año, S=mes, T=Marca 2, V=Equipo Asociado (familia).
    """
    MESES_ABR = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
                 "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    ws = None
    for name in wb.sheetnames:
        n = name.strip().lower()
        if "repuesto" in n and "vend" in n:
            ws = wb[name]
            break
    if ws is None:
        return {}

    # celdas[(marca, anio, mes)] = [monto, cantidad]
    celdas = defaultdict(lambda: [0.0, 0.0])
    # clientes[marca][cliente][anio] = [monto, cantidad]. Se guarda abierto por
    # año para que el panel pueda recalcular el top de clientes según el
    # segmentador 2025 / 2026 / ambos sin volver a leer el Excel.
    clientes = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: [0.0, 0.0])))
    # cli_mes[(cliente, anio, mes)] = [monto, cantidad] — serie mensual por
    # cliente, para el gráfico de evolución con selector de cliente.
    cli_mes   = defaultdict(lambda: [0.0, 0.0])
    cli_marca = defaultdict(lambda: defaultdict(float))
    # cli_mm[(cliente, marca, anio, mes)] = [monto, cantidad] — desglose por
    # marca dentro de cada mes, para las barras apiladas del gráfico y su
    # tooltip. Sólo se guardan los pares (cliente, marca) con venta real.
    cli_mm    = defaultdict(lambda: [0.0, 0.0])
    # Familia de producto = "Equipo Asociado" (col V): AUTOCLAVES, LAVADORA,
    # OTROS. Se abre por marca y por cliente para poder segmentar las tablas.
    fam_mes   = defaultdict(lambda: [0.0, 0.0])   # (familia, anio, mes)
    marca_fam = defaultdict(lambda: [0.0, 0.0])   # (marca, familia, anio, mes)
    cli_fam   = defaultdict(lambda: [0.0, 0.0])   # (cliente, familia, anio, mes)
    familias  = set()
    periodos = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 20:
            continue
        marca = safe_str(row[19]).strip().upper()
        if not marca:
            continue
        try:
            mes = int(row[18])
        except (TypeError, ValueError):
            continue
        if not (1 <= mes <= 12):
            continue
        anio = safe_str(row[17]).strip()
        if not anio:
            continue
        monto = to_float(row[15])
        cant  = to_float(row[12])
        cli   = safe_str(row[2]).strip() or "(sin cliente)"

        celdas[(marca, anio, mes)][0] += monto
        celdas[(marca, anio, mes)][1] += cant
        clientes[marca][cli][anio][0] += monto
        clientes[marca][cli][anio][1] += cant
        cli_mes[(cli, anio, mes)][0] += monto
        cli_mes[(cli, anio, mes)][1] += cant
        cli_marca[cli][marca] += monto
        cli_mm[(cli, marca, anio, mes)][0] += monto
        cli_mm[(cli, marca, anio, mes)][1] += cant
        fam = safe_str(row[21]).strip().upper() if len(row) > 21 else ""
        fam = fam or "SIN CLASIFICAR"
        familias.add(fam)
        fam_mes[(fam, anio, mes)][0] += monto
        fam_mes[(fam, anio, mes)][1] += cant
        marca_fam[(marca, fam, anio, mes)][0] += monto
        marca_fam[(marca, fam, anio, mes)][1] += cant
        cli_fam[(cli, fam, anio, mes)][0] += monto
        cli_fam[(cli, fam, anio, mes)][1] += cant
        periodos.add((anio, mes))

    if not celdas:
        return {}

    per = sorted(periodos, key=lambda p: (int(p[0]), p[1]))
    meses = [{"a": a, "m": m, "lbl": f"{MESES_ABR[m-1]} {a[2:]}"} for a, m in per]
    n = len(per)
    idx = {p: i for i, p in enumerate(per)}

    anios = sorted({a for a, _ in per})
    marcas = sorted({m for (m, _, _) in celdas})
    out = {}
    for marca in marcas:
        monto = [0.0] * n
        cant  = [0.0] * n
        for (mk, a, m), (v, q) in celdas.items():
            if mk != marca:
                continue
            i = idx[(a, m)]
            monto[i] += v
            cant[i]  += q
        # Lista completa de clientes abierta por año: el panel arma el top
        # según el año seleccionado.
        cli_out = []
        for c, por_anio in clientes[marca].items():
            reg = {"c": c}
            for a in anios:
                v, q = por_anio.get(a, [0.0, 0.0])
                reg["m" + a[2:]] = round(v)
                reg["q" + a[2:]] = round(q)
            cli_out.append(reg)
        cli_out.sort(key=lambda r: -sum(r[k] for k in r if k.startswith("m")))
        # Serie mensual de esta marca abierta por familia de producto
        fam_d = {}
        for (mk, fm, a, m), (v, q) in marca_fam.items():
            if mk != marca:
                continue
            e = fam_d.setdefault(fm, {"m": [0.0] * n, "q": [0.0] * n})
            i = idx[(a, m)]
            e["m"][i] += v
            e["q"][i] += q
        out[marca] = {
            "monto":      [round(x) for x in monto],
            "cant":       [round(x) for x in cant],
            "monto_tot":  round(sum(monto)),
            "cant_tot":   round(sum(cant)),
            "n_clientes": len(cli_out),
            "clientes":   cli_out,
            "fam":        {k: {"m": [round(x) for x in v["m"]],
                               "q": [round(x) for x in v["q"]]}
                           for k, v in fam_d.items()},
        }

    marcas_sorted = sorted(out, key=lambda m: -out[m]["monto_tot"])
    tot_monto = [round(sum(out[m]["monto"][i] for m in out)) for i in range(n)]
    tot_cant  = [round(sum(out[m]["cant"][i]  for m in out)) for i in range(n)]
    todos_cli = set()
    for marca in clientes:
        todos_cli.update(clientes[marca].keys())

    # Serie mensual por cliente (para el gráfico con selector de cliente)
    series_mes = defaultdict(lambda: [[0.0] * n, [0.0] * n])
    for (ck, a, m), (v, q) in cli_mes.items():
        i = idx[(a, m)]
        series_mes[ck][0][i] += v
        series_mes[ck][1][i] += q

    # Desglose mensual por marca dentro de cada cliente
    det = defaultdict(lambda: defaultdict(lambda: [[0.0] * n, [0.0] * n]))
    for (ck, mk_, a, m), (v, q) in cli_mm.items():
        i = idx[(a, m)]
        det[ck][mk_][0][i] += v
        det[ck][mk_][1][i] += q

    # Desglose mensual por familia dentro de cada cliente
    detf = defaultdict(lambda: defaultdict(lambda: [[0.0] * n, [0.0] * n]))
    for (ck, fm, a, m), (v, q) in cli_fam.items():
        i = idx[(a, m)]
        detf[ck][fm][0][i] += v
        detf[ck][fm][1][i] += q

    cli_out = {}
    for c in todos_cli:
        mo, qt = series_mes[c]
        mk = sorted(cli_marca[c].items(), key=lambda x: -x[1])
        cli_out[c] = {
            "monto":     [round(x) for x in mo],
            "cant":      [round(x) for x in qt],
            "monto_tot": round(sum(mo)),
            "cant_tot":  round(sum(qt)),
            "marcas":    [[k, round(v)] for k, v in mk],
            "n_marcas":  len(mk),
            # {marca: {"m": [n montos], "q": [n cantidades]}} ordenado por monto
            "det": {
                k: {"m": [round(x) for x in det[c][k][0]],
                    "q": [round(x) for x in det[c][k][1]]}
                for k, _ in mk
            },
            "fam": {
                k: {"m": [round(x) for x in v[0]],
                    "q": [round(x) for x in v[1]]}
                for k, v in detf[c].items()
            },
        }

    # Resumen por familia de producto: serie mensual y apertura por marca
    fam_out = {}
    for fm in familias:
        mo = [0.0] * n
        qt = [0.0] * n
        for (f2, a, m), (v, q) in fam_mes.items():
            if f2 != fm:
                continue
            i = idx[(a, m)]
            mo[i] += v
            qt[i] += q
        por_marca = defaultdict(lambda: [0.0, 0.0])
        for (mk, f2, a, m), (v, q) in marca_fam.items():
            if f2 != fm:
                continue
            por_marca[mk][0] += v
            por_marca[mk][1] += q
        fam_out[fm] = {
            "monto":     [round(x) for x in mo],
            "cant":      [round(x) for x in qt],
            "monto_tot": round(sum(mo)),
            "cant_tot":  round(sum(qt)),
            "marcas":    [{"k": k, "monto": round(v[0]), "cant": round(v[1])}
                          for k, v in sorted(por_marca.items(), key=lambda x: -x[1][0])],
        }
    familias_sorted = sorted(fam_out, key=lambda f: -fam_out[f]["monto_tot"])

    return {
        "meses":       meses,
        "anios":       anios,
        "marcas":      marcas_sorted,
        "familias":    familias_sorted,
        "fam":         fam_out,
        "data":        out,
        "tot_monto":   tot_monto,
        "tot_cant":    tot_cant,
        "tot_monto_g": sum(tot_monto),
        "tot_cant_g":  sum(tot_cant),
        "n_clientes":  len(todos_cli),
        "cli_serie":   cli_out,
    }


def read_pipeline_st(wb):
    """Pipeline comercial de equipos, para el potencial de ST por garantías.

    Fuente: las hojas "PIPELINE Esterilización", "PIPELINE Dental" y
    "PIPELINE Endoscopía". Las tres tienen la misma estructura:
      A=Nombre Cliente  B=Productos  C=Probabilidad de venta (decimal 0–1)
      D=Monto Negocio   E=Mes probable de facturación  F=Nombre Analisis
      G=Año

    Sólo se leen las filas crudas; el porcentaje de garantía por línea y el
    margen se aplican en hoja_pipeline.js, igual que las tarifas UF de la
    Base Instalada, para que queden a la vista y sean fáciles de ajustar.
    """
    MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    MES_IDX = {_norm_cli(m): i + 1 for i, m in enumerate(MESES)}
    HOJAS = [("esteriliz", "Esterilización"), ("dental", "Dental"), ("endoscop", "Endoscopía")]

    items, lineas, anios, meses_vistos = [], [], set(), set()
    for clave, linea in HOJAS:
        ws = None
        for name in wb.sheetnames:
            n = _norm_cli(name)
            if n.startswith("PIPELINE") and _norm_cli(clave) in n:
                ws = wb[name]
                break
        if ws is None:
            print(f"  ADVERTENCIA: no se encontró la hoja PIPELINE de {linea}.")
            continue
        lineas.append(linea)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 7:
                continue
            cli = safe_str(row[0]).strip()
            monto = to_float(row[3])
            if not cli and monto == 0:
                continue
            anio = safe_str(row[6]).strip()
            try:
                anio = str(int(float(anio)))
            except (TypeError, ValueError):
                anio = anio or "Sin año"
            mes_txt = safe_str(row[4]).strip()
            mes_n = MES_IDX.get(_norm_cli(mes_txt), 0)
            if mes_n:
                mes_txt = MESES[mes_n - 1]
            else:
                mes_txt = mes_txt or "Sin definir"
            na = safe_str(row[5]).strip()
            if na in ("0", "0.0"):
                na = ""
            # La probabilidad viene como decimal con ruido de coma flotante
            # (0.30000000000000004); se redondea a punto porcentual.
            prob = round(to_float(row[2]) * 100)
            anios.add(anio)
            meses_vistos.add(mes_txt)
            items.append({
                "cli":    cli or "(sin cliente)",
                "na":     na,
                "prod":   safe_str(row[1]).strip() or "—",
                "prob":   max(0, min(100, prob)),
                "monto":  round(monto),
                "mes":    mes_txt,
                "mes_n":  mes_n,
                "anio":   anio,
                "linea":  linea,
            })

    if not items:
        return {}
    items.sort(key=lambda x: -x["monto"])
    return {
        "lineas": lineas,
        "anios":  sorted(anios),
        "meses":  MESES,
        "items":  items,
        "n":      len(items),
        "n_clientes": len({_norm_cli(i["cli"]) for i in items}),
        "monto_tot":  round(sum(i["monto"] for i in items)),
    }


def read_gd_costos(wb):
    """Costo de repuestos consumidos por cliente.

    Fuente: hoja "GD" (guías de despacho y traslados de bodega, una fila por
    movimiento). Col J = Nombre Cliente, col R = Costo Total. Se agrupa por
    cliente normalizado; es el costo que la hoja Panel Fact Cliente resta de
    lo facturado para obtener el margen bruto.

    Reemplaza a "Total Costo Linea" (col Q de la BBDD de facturación), que
    sólo recogía el costo de las líneas facturadas y dejaba a la mitad de los
    clientes en cero.
    """
    ws = next((wb[n] for n in wb.sheetnames if n.strip().upper() == "GD"), None)
    if ws is None:
        print("  ADVERTENCIA: no se encontró la hoja 'GD'; el costo por cliente queda en cero.")
        return {}
    idx = {}
    n_filas = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 18:
            continue
        cli = safe_str(row[9]).strip()
        if not cli:
            continue
        k = _norm_cli(cli)
        e = idx.setdefault(k, {"nombre": cli, "costo": 0.0})
        e["costo"] += to_float(row[17])
        n_filas += 1
    print(f"       GD: {n_filas} movimientos | {len(idx)} clientes | "
          f"costo MM${sum(e['costo'] for e in idx.values())/1e6:,.1f}")
    for e in idx.values():
        e["costo"] = round(e["costo"])
    return idx


def read_back_order_idx(wb):
    """Unidades pedidas al proveedor por SKU.

    Hoja "Back Order": col G = SKU, col K = Cantidad Solicitada. Se suman
    todas las filas sin mirar "Estatus OC" — ese campo indica si la orden
    está pagada, no si la mercadería llegó, así que no hay doble conteo
    contra el stock de bodega.
    """
    idx = defaultdict(float)
    ws = next((wb[n] for n in wb.sheetnames if n.strip().lower() == "back order"), None)
    if ws is None:
        return {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 11:
            continue
        sku = safe_str(row[6]).strip()
        if sku:
            idx[sku] += to_float(row[10])
    return {k: round(v, 2) for k, v in idx.items() if v}


def read_clientes_relevantes(wb):
    """Venta de repuestos por cliente y SKU en ventanas móviles de 6 y 12 meses.

    Alimenta la hoja "Clientes Relevantes": el top de clientes por compra de
    repuestos y la tabla de fill rate, que dimensiona el stock objetivo a
    partir de lo que cada cliente consumió en el período.

    Fuente: hoja "Repuestos Vendidas". Columnas C=cliente, M=cantidad,
    N=nombre producto, P=precio de venta, R=año, S=mes, T=Marca 2,
    U=SKU 2 (normalizado, viene lleno en todas las filas), V=Equipo Asociado.

    Las ventanas se cuentan hacia atrás desde el último mes con datos, no
    desde la fecha de hoy: si el Excel se actualiza con rezago, "últimos 6
    meses" sigue significando los 6 meses efectivamente cargados.
    """
    MESES_ABR = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
                 "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    ws = None
    for name in wb.sheetnames:
        n = name.strip().lower()
        if "repuesto" in n and "vend" in n:
            ws = wb[name]
            break
    if ws is None:
        return {}

    filas = []
    periodos = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 21:
            continue
        try:
            mes = int(row[18])
        except (TypeError, ValueError):
            continue
        if not (1 <= mes <= 12):
            continue
        try:
            anio = int(safe_str(row[17]).strip())
        except (TypeError, ValueError):
            continue
        sku = safe_str(row[20]).strip() or safe_str(row[11]).strip()
        if not sku:
            continue
        periodos.add((anio, mes))
        filas.append((
            safe_str(row[2]).strip() or "(sin cliente)",
            sku,
            safe_str(row[13]).strip(),
            safe_str(row[19]).strip().upper() or "SIN MARCA",
            safe_str(row[21]).strip().upper() if len(row) > 21 else "SIN CLASIFICAR",
            anio, mes,
            to_float(row[12]),
            to_float(row[15]),
        ))
    if not filas:
        return {}

    per = sorted(periodos)
    ult = per[-12:]                       # ventana larga = base de todo
    idx = {p: i for i, p in enumerate(ult)}
    n = len(ult)
    corte6 = max(0, n - 6)                # índice desde el que empieza la ventana corta
    lbl = lambda p: MESES_ABR[p[1] - 1] + " " + str(p[0])[2:]

    # agg[(cliente, sku)] = [prod, marca, familia, [cant x n], [monto x n]]
    agg = {}
    for cli, sku, prod, marca, fam, a, m, q, v in filas:
        if (a, m) not in idx:
            continue
        i = idx[(a, m)]
        k = (cli, sku)
        e = agg.get(k)
        if e is None:
            e = agg[k] = [prod, marca, fam, [0.0] * n, [0.0] * n]
        if prod and not e[0]:
            e[0] = prod
        e[3][i] += q
        e[4][i] += v

    # Vuelca a estructura por cliente, con los totales de cada ventana ya hechos
    cli = {}
    for (c, sku), (prod, marca, fam, qs, vs) in agg.items():
        d = cli.get(c)
        if d is None:
            d = cli[c] = {"cliente": c, "skus": [], "serie_m": [0.0] * n, "serie_q": [0.0] * n}
        q12, v12 = sum(qs), sum(vs)
        q6,  v6  = sum(qs[corte6:]), sum(vs[corte6:])
        d["skus"].append({
            "sku":   sku,
            "prod":  prod or sku,
            "marca": marca,
            "fam":   fam,
            "q6":  round(q6, 2),  "v6":  round(v6),
            "q12": round(q12, 2), "v12": round(v12),
        })
        for i in range(n):
            d["serie_m"][i] += vs[i]
            d["serie_q"][i] += qs[i]

    clientes = []
    for c, d in cli.items():
        s6  = [x for x in d["skus"] if x["q6"] > 0 or x["v6"] > 0]
        s12 = d["skus"]
        d["skus"].sort(key=lambda x: -x["v12"])
        clientes.append({
            "cliente": c,
            "m6":  round(sum(x["v6"]  for x in s12)), "q6":  round(sum(x["q6"]  for x in s12), 2),
            "m12": round(sum(x["v12"] for x in s12)), "q12": round(sum(x["q12"] for x in s12), 2),
            "n6":  len(s6), "n12": len(s12),
            "serie_m": [round(x) for x in d["serie_m"]],
            "serie_q": [round(x, 2) for x in d["serie_q"]],
            "skus": s12,
        })
    clientes.sort(key=lambda x: -x["m12"])

    tot = lambda k: round(sum(c[k] for c in clientes), 2)
    return {
        "meses":     [lbl(p) for p in ult],
        "corte6":    corte6,
        "periodo6":  lbl(ult[corte6]) + " – " + lbl(ult[-1]),
        "periodo12": lbl(ult[0]) + " – " + lbl(ult[-1]),
        "clientes":  clientes,
        "tot_m6":  tot("m6"),  "tot_q6":  tot("q6"),
        "tot_m12": tot("m12"), "tot_q12": tot("q12"),
        "n_clientes": len(clientes),
    }


def read_inventario_ts(wb):
    """Inventario de repuestos en bodega, agregado por marca y SKU.

    Fuente: hoja "Inventario Bodega" (una fila por SKU y bodega). Replica la
    tabla dinámica "TD Inventario TS": agrupa por (marca, SKU) sumando stock y
    costo total, y promediando el costo unitario — 82 SKUs aparecen en más de
    una bodega, por eso el promedio y no el valor de una fila cualquiera.

    Columnas: C=SKU, E=Descripción, G=Categoría, H=FirmName (marca),
              L=En stock, M=Costo del artículo, N=Costo Total.
    """
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == "inventario bodega":
            ws = wb[name]
            break
    if ws is None:
        return {}

    # {marca: {sku: {"d": desc, "st": stock, "cus": [costos unit], "ct": costo total}}}
    # El costo unitario se promedia sólo a nivel SKU (82 SKU están en más de una
    # bodega). A nivel marca NO se expone un promedio: cualquier forma de
    # agregarlo es engañosa — el promedio simple pesa igual un tornillo de $47
    # que un generador de $3,8 MM, y ponderarlo por costo eleva al cuadrado los
    # SKU caros. La tabla muestra sólo stock y costo total por marca.
    marcas = defaultdict(lambda: defaultdict(lambda: {"d": "", "st": 0.0, "cus": [], "ct": 0.0}))

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 14:
            continue
        marca = safe_str(row[7]).strip()
        if not marca:
            continue
        sku = safe_str(row[2]).strip() or "(sin SKU)"
        it  = marcas[marca][sku]
        if not it["d"]:
            it["d"] = safe_str(row[4]).strip()
        it["st"] += to_float(row[11])
        it["ct"] += to_float(row[13])
        cu = row[12]
        if isinstance(cu, (int, float)):
            it["cus"].append(float(cu))


    out       = {}
    tot_stock = 0.0
    tot_costo = 0.0
    tot_skus  = 0
    for marca, skus in marcas.items():
        items = []
        m_st  = 0.0
        m_ct  = 0.0
        for sku, it in skus.items():
            cu = sum(it["cus"]) / len(it["cus"]) if it["cus"] else 0.0
            items.append({
                "sku": sku,
                "d":   it["d"],
                "st":  round(it["st"], 2),
                "cu":  round(cu),
                "ct":  round(it["ct"]),
            })
            m_st += it["st"]
            m_ct += it["ct"]
        items.sort(key=lambda x: -x["ct"])
        out[marca] = {
            "stock":   round(m_st, 2),
            "ct":      round(m_ct),
            "n_skus":  len(items),
            "items":   items,
        }
        tot_stock += m_st
        tot_costo += m_ct
        tot_skus  += len(items)

    marcas_sorted = sorted(out, key=lambda m: -out[m]["ct"])
    return {
        "marcas":      marcas_sorted,
        "data":        out,
        "total_stock": round(tot_stock, 2),
        "total_costo": round(tot_costo),
        "total_skus":  tot_skus,
        "n_marcas":    len(out),
    }


def read_costo_tecnicos(wb, n_meses):
    """Costo de personal técnico dentro del Costo de Ventas.

    Fuente: hoja "EERR S&S", bloque bajo "Presupuesto acumulado al periodo"
    (fila "Costo Empresas Técnicos"), con 4 columnas por mes igual que la
    sección EERR principal (Real = col 2 + i*4). Sólo trae el Real; el Excel
    no tiene un presupuesto para esta apertura de "Costo de ventas" en
    técnicos/personal vs. repuestos y otros, así que el resto ("Costo
    Repuestos y Otros") se calcula como el residuo contra "Costo de ventas"
    y también queda sólo en Real.
    """
    ws = next((wb[n] for n in wb.sheetnames if n.strip().upper() == "EERR S&S"), None)
    if ws is None:
        return [0.0] * n_meses
    fila = None
    for row in ws.iter_rows(min_row=1, max_row=80, values_only=True):
        if len(row) > 1 and "TECNIC" in _norm_cli(row[1]) and "COSTO" in _norm_cli(row[1]):
            fila = row
            break
    if fila is None:
        return [0.0] * n_meses
    out = []
    for i in range(n_meses):
        col = 2 + i * 4
        v = to_float(fila[col]) if col < len(fila) else 0.0
        out.append(round(-abs(v) / 1e6, 3))
    return out


def read_ratios2(wb):
    ws = None
    for name in wb.sheetnames:
        if "ratio" in name.lower() and "2" in name:
            ws = wb[name]
            break
    if ws is None:
        return {}

    # Estructura: col B = etiqueta, Real de mes i = col (2 + i*4), 0-indexed
    # Fila 7=Ingresos, 8=Contratos, 9=Otras, 10=CdV, 11=Margen, 12=Margen%,
    # 14=Empleados, 15=Otros, 17=EBITDA Directo, 20=GAV Indirecto, 30=EBITDA Empresa
    MAX_MONTHS = 12
    MAX_COL    = max(2 + MAX_MONTHS * 4 + 1, 2 + MAX_MONTHS * 2 + 1)  # cubre formato EERR y RATIOS

    rows = list(ws.iter_rows(min_row=1, max_row=70, max_col=MAX_COL, values_only=True))

    MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
             "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

    def get_val(row_idx, col_idx):
        row = rows[row_idx] if row_idx < len(rows) else []
        if col_idx is not None and col_idx < len(row) and row[col_idx] is not None:
            try: return float(row[col_idx])
            except (TypeError, ValueError): return 0.0
        return 0.0

    # Los meses se resuelven leyendo las dos filas de encabezado (fila 5 = tipo
    # de columna, fila 6 = mes) en vez de asumir 4 columnas por mes. La hoja no
    # es regular: julio quedó con 3 columnas (sin "Variación Ptto %"), y con
    # paso fijo el bloque "Periodo Actual" del final entraba como un mes más.
    def _mes_de(v):
        if isinstance(v, (datetime, date)):
            return (v.year, v.month)
        t = safe_str(v).strip()
        m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})$", t)
        if m:
            a = int(m.group(3))
            return (a + 2000 if a < 100 else a, int(m.group(2)))
        return None

    TIPOS = {"REAL": "real", "PTTO": "ptto",
             "VARIACION PTTO": "var", "VARIACION PTTO %": "vp"}
    fila_tipo = rows[4] if len(rows) > 4 else []
    fila_mes  = rows[5] if len(rows) > 5 else []
    meses_cols, orden = {}, []
    for ci in range(2, min(len(fila_mes), MAX_COL)):
        clave = _mes_de(fila_mes[ci])
        if clave is None:
            continue                      # "Periodo Actual", acumulados, etc.
        tipo = TIPOS.get(_norm_cli(fila_tipo[ci]) if ci < len(fila_tipo) else "")
        if tipo is None:
            continue
        if clave not in meses_cols:
            meses_cols[clave] = {}
            orden.append(clave)
        meses_cols[clave].setdefault(tipo, ci)
    orden.sort()

    def _col(month_i, tipo):
        if month_i >= len(orden):
            return None
        return meses_cols[orden[month_i]].get(tipo)

    def get_real(row_idx, month_i):   return get_val(row_idx, _col(month_i, "real"))
    def get_ptto(row_idx, month_i):   return get_val(row_idx, _col(month_i, "ptto"))
    def get_var(row_idx, month_i):    return get_val(row_idx, _col(month_i, "var"))
    def get_varpct(row_idx, month_i): return get_val(row_idx, _col(month_i, "vp"))

    def arr(row_idx, n):    return [round(get_real(row_idx, i),   3) for i in range(n)]
    def arrp(row_idx, n):   return [round(get_ptto(row_idx, i),   3) for i in range(n)]
    def arrv(row_idx, n):   return [round(get_var(row_idx, i),    3) for i in range(n)]
    def arrvp(row_idx, n):  return [round(get_varpct(row_idx, i), 4) for i in range(n)]

    # Las filas se ubican por su etiqueta en la columna B y no por un índice
    # fijo: la hoja ya se ha reordenado dos veces (se insertaron "Costo
    # Tecnicos y Personal" y "Costos Repuestos y Otros" bajo Costo de ventas,
    # corriendo todo lo de abajo) y con índices fijos el EERR quedaba
    # silenciosamente descuadrado. Se toma la primera fila que contenga el
    # fragmento, por eso "% EBITDA" resuelve al de EBITDA Directo, que es el
    # que va primero.
    ETIQUETAS = [
        ("ingresos_totales",        "INGRESOS DE ACTIVIDADES"),
        ("ingresos_contratos",      "INGRESOS POR CONTRATOS"),
        ("ingresos_otras",          "INGRESOS POR OTRAS ACTIVIDADES"),
        ("costo_ventas",            "COSTO DE VENTAS"),
        ("costo_tecnicos",          "COSTO TECNICOS"),
        ("costo_repuestos",         "REPUESTOS Y OTROS"),
        ("margen_mm",               "MARGEN DEL PRODUCTO"),
        ("margen_pct",              "% MARGEN"),
        ("gastos_empleados",        "BENEFICIOS A LOS EMPLEADOS"),
        ("otros_gastos",            "OTROS GASTOS POR NATURALEZA"),
        ("ebitda_directo",          "EBITDA DIRECTO"),
        ("ebitda_directo_pct",      "% EBITDA"),
        ("gav_indirecto",           "GAV INDIRECTO"),
        ("ebitda_indirecto",        "EBITDA INDIRECTO"),
        ("finiquitos",              "FINIQUITOS"),
        ("multas",                  "MULTAS"),
        ("prov_obsolescencias",     "OBSOLESCENCIAS"),
        ("prov_incobrables",        "INCOBRA"),
        ("prov_habilitacion",       "HABILITACION"),
        ("total_gastos_adicionales","TOTAL GASTOS ADICIONALES"),
        ("ebitda_empresa",          "EBITDA EMPRESA"),
        ("depreciacion",            "DEPRECIACION"),
        ("resultado_operacional",   "RESULTADO OPERACIONAL"),
        ("otros_ingresos_funcion",  "OTROS INGRESOS POR FUNCION"),
        ("ingreso_financiero",      "INGRESO FINANCIERO"),
        ("costo_financiero",        "COSTO FINANCIERO"),
        ("otros_gastos_funcion",    "OTROS GASTOS POR FUNCION"),
        ("diferencia_cambio",       "DIFERENCIA DE CAMBIO"),
        ("resultado_no_operacional","RESULTADO NO OPERACIONAL"),
        ("resultado_antes_imp",     "RESULTADO ANTES DE IMPUESTOS"),
        ("impuesto_renta",          "IMPUESTO A LA RENTA"),
        ("resultado_ejercicio",     "RESULTADO DEL EJERCICIO"),
    ]
    PORCENTAJES = {"margen_pct", "ebitda_directo_pct"}

    idx_fila = {}
    for ri, row in enumerate(rows):
        etq = _norm_cli(row[1]) if len(row) > 1 and row[1] is not None else ""
        if not etq:
            continue
        for clave, frag in ETIQUETAS:
            if clave not in idx_fila and frag in etq:
                idx_fila[clave] = ri
                break
    faltan = [c for c, _ in ETIQUETAS if c not in idx_fila]
    if faltan:
        print(f"  ADVERTENCIA: filas no encontradas en Ratio Costos 2: {', '.join(faltan)}")

    # Detectar mes_cierre: meses consecutivos con Ingresos != 0.
    # Paramos en el primer cero para no capturar la columna Total al final.
    _ri_ing = idx_fila.get("ingresos_totales", 6)
    mes_cierre = 0
    for i in range(len(orden)):
        if abs(get_real(_ri_ing, i)) > 0.001:
            mes_cierre = i + 1
        else:
            break
    if mes_cierre == 0:
        return {}

    n = mes_cierre
    pct  = lambda idx: [round(get_real(idx, i),   4) for i in range(n)]
    pctp = lambda idx: [round(get_ptto(idx, i),   4) for i in range(n)]
    pctv = lambda idx: [round(get_var(idx, i),    4) for i in range(n)]
    pctvp= lambda idx: [round(get_varpct(idx, i), 4) for i in range(n)]

    def row4(ri):  # (real, ptto, var, varpct) para filas MM$
        return arr(ri,n), arrp(ri,n), arrv(ri,n), arrvp(ri,n)
    def pct4(ri):  # idem para filas %
        return pct(ri), pctp(ri), pctv(ri), pctvp(ri)

    D = {}
    for clave, _ in ETIQUETAS:
        ri = idx_fila.get(clave)
        if ri is None:
            D[clave] = ([0.0] * n, [0.0] * n, [0.0] * n, [0.0] * n)
        elif clave in PORCENTAJES:
            D[clave] = pct4(ri)
        else:
            D[clave] = row4(ri)

    def col(clave, k=0):
        return D[clave][k]

    # La apertura del costo de ventas viene sólo en Real; el Excel no la
    # presupuesta. Si la hoja no trae las dos líneas nuevas, "Repuestos y
    # Otros" se calcula como el residuo contra el costo de ventas.
    costo_tecnicos  = col("costo_tecnicos")
    costo_repuestos = col("costo_repuestos")
    if not any(costo_repuestos):
        costo_repuestos = [round(col("costo_ventas")[i] - costo_tecnicos[i], 3) for i in range(n)]

    salida = {
        "mes_cierre": mes_cierre,
        "meses":      [MESES[orden[i][1] - 1] for i in range(n)],
        "costo_tecnicos":  costo_tecnicos,
        "costo_repuestos": costo_repuestos,
    }
    for clave, _ in ETIQUETAS:
        if clave in ("costo_tecnicos", "costo_repuestos"):
            continue
        r, pp, vv, vp = D[clave]
        salida[clave]           = r
        salida[clave + "_p"]    = pp
        salida[clave + "_v"]    = vv
        salida[clave + "_vp"]   = vp
    return salida


def _read_ratios_section(rows, MESES):
    """Lee la sección RATIOS (filas 49-64) de Ratio Costos 2.
    Formato: 2 columnas por mes (Real = col[2+i*2], PTTO = col[3+i*2]).
    Filas clave (índice 0-based desde fila Excel 1):
      52 → Ingresos ordinarios, 53 → Contratos, 54 → Otras actividades
      55 → Costo Total, 56 → Costo de ventas
      58 → Empleados directos, 59 → Otros directos
      60 → GAV Indirecto, 61 → GAV Total, 63 → Margen del Producto
    """
    def gv(row_idx, col_idx):
        row = rows[row_idx] if row_idx < len(rows) else []
        if col_idx < len(row) and row[col_idx] is not None:
            try: return float(row[col_idx])
            except: return 0.0
        return 0.0

    def real(ri, i): return round(gv(ri, 2 + i * 2), 3)
    def ptto(ri, i): return round(gv(ri, 3 + i * 2), 3)

    # Detectar meses con datos (fila 53 = index 52, Ingresos)
    n_r = 0
    for i in range(12):
        if abs(gv(52, 2 + i * 2)) > 0.01:
            n_r = i + 1
        else:
            break

    if n_r == 0:
        return {"ratios_kpis": {}}

    def arr_r(ri): return [real(ri, i) for i in range(n_r)]
    def arr_p(ri): return [ptto(ri, i) for i in range(n_r)]

    return {
        "ratios_kpis": {
            "n": n_r,
            "meses": MESES[:n_r],
            "ingresos_r":     arr_r(52), "ingresos_p":     arr_p(52),   # fila 53
            "contratos_r":    arr_r(53), "contratos_p":    arr_p(53),   # fila 54
            "otras_r":        arr_r(54), "otras_p":        arr_p(54),   # fila 55
            "costo_total_r":  arr_r(55), "costo_total_p":  arr_p(55),   # fila 56
            "cdv_r":          arr_r(56), "cdv_p":          arr_p(56),   # fila 57
            "empleados_r":    arr_r(58), "empleados_p":    arr_p(58),   # fila 59
            "otros_dir_r":    arr_r(59), "otros_dir_p":    arr_p(59),   # fila 60
            "gav_ind_r":      arr_r(60), "gav_ind_p":      arr_p(60),   # fila 61
            "gav_total_r":    arr_r(61), "gav_total_p":    arr_p(61),   # fila 62
            "margen_prod_r":  arr_r(63), "margen_prod_p":  arr_p(63),   # fila 64
        }
    }


def read_resumen_tipos_programas(wb):
    ws = wb['Resumen Tipos Programas']
    rows_out = []
    for row in ws.iter_rows(min_row=3, max_row=7, values_only=True):
        prog = safe_str(row[1]) if row[1] else ""
        if not prog:
            continue
        rows_out.append({
            "programa":        prog,
            "clientes":        round(to_float(row[2])),
            "contratos":       to_int(row[3]),
            "fac_promedio":    round(to_float(row[4])),
            "fac_esperada":    round(to_float(row[5])),
            "pct_cartera":     round(to_float(row[6]), 4) if row[6] is not None else None,
            "pct_margen":      round(to_float(row[7]), 4) if row[7] is not None else None,
            "margen_total":    round(to_float(row[8])),
            "margen_promedio": round(to_float(row[9])),
            "dur_promedio":    round(to_float(row[10]), 1),
            "vig_promedio":    round(to_float(row[11]), 1),
        })
    return rows_out


def completar_mapa(mapa_data, fact_clientes):
    """Agrega al mapa los clientes que facturan pero no están en BASE MAPA.

    El mapa se arma desde la hoja BASE MAPA, que no cubre a todos los clientes
    con facturación: los que faltan hacían que el total del mapa quedara por
    debajo del de las demás hojas. En vez de dejar el descuadre, se agregan con
    su facturación y sin coordenadas —el panel ya sabe dibujar filas sin lat/lon,
    aparecen en las tablas pero no como punto—, de modo que el total del mapa
    siempre cuadre con la facturación real. Cuando el cliente se cargue en BASE
    MAPA con su ubicación, esta función deja de agregarlo sola.
    """
    if not mapa_data or not fact_clientes:
        return mapa_data
    presentes = {_norm_cli(c["n"]) for c in mapa_data}
    faltan = [c for c in fact_clientes
              if c.get("real") and _norm_cli(c["cliente"]) not in presentes]
    for c in faltan:
        mapa_data.append({
            "n":            c["cliente"],
            "tipo":         c.get("tipo_cli") or "",
            "ingreso_2025": 0,
            "ingreso_2026": int(round(c["real"])),
            "ingreso":      int(round(c["real"])),
            "bi": 0, "eq": {}, "region": "Sin región", "comuna": "",
            "contratos": 0, "pipe": 0,
            "lat": None, "lon": None,     # sin ubicación: no se dibuja en el mapa
            "cc": 0, "margen": 0, "sat": None,
            "pot_eq": 0, "pot_eq_ester": 0, "pot_eq_endo": 0, "pot_eq_dental": 0,
            "pot_st": 0, "pot_st_gar": 0, "pot_st_contr": 0, "pot": 0,
            "sin_base_mapa": True,        # para poder distinguirlos en el panel
        })
    if faltan:
        tot = sum(c["real"] for c in faltan)
        print(f"       MAPA: +{len(faltan)} clientes facturados sin fila en BASE MAPA "
              f"(MM${tot/1e6:,.1f}) agregados sin ubicacion")
    return mapa_data


def enrich_mapa_data(mapa_data, contratos, satisf):
    """Agrega conteo de contratos activos y satisfacción. cc ya viene del Excel."""
    active_by_cli = {}
    for c in contratos:
        if c["estado"] != "Activado":
            continue
        cli = c["cliente"].upper().strip()
        active_by_cli[cli] = active_by_cli.get(cli, 0) + 1

    # Satisfacción: calidad 0–7 → escala 1–5
    sat_by_inst = {}
    for inst in satisf.get("instituciones", []):
        key = inst["institucion"].upper().strip()
        cal = inst.get("calidad", 0)
        if cal and cal > 0:
            sat_by_inst[key] = round(1 + (cal / 7) * 4, 1)

    for c in mapa_data:
        cli = c["n"].upper().strip()
        c["contratos"] = active_by_cli.get(cli, 0)
        # cc ya viene de col AA — no se sobrescribe
        # sat: si col AB tiene valor lo respeta; si es None usa el cruce con SATISFACCION
        if c["sat"] is None:
            c["sat"] = sat_by_inst.get(cli, None)

    return mapa_data


# ══════════════════════════════════════════════════════════════════════════════
# 6b. DESGLOSE DE INGRESOS — Ingreso por Contratos (por línea de negocio) +
#     Otros Ingresos (Trazabilidad / REAS / Marca). Reconciliado 1:1 con la
#     facturación real mensual (fac_arr) y con contr_real_monthly.
# ══════════════════════════════════════════════════════════════════════════════
def compute_desglose_ingresos(contratos, panel_raw, bbdd, contr_real_monthly, fac_arr, mes_corte, mapa_data=None):
    n = mes_corte
    if n <= 0:
        return {}

    contratos_by_cli = defaultdict(list)
    for c in contratos:
        contratos_by_cli[c["cliente"]].append(c)

    LINEAS = ["Esterilización", "Endoscopía", "Dental"]
    linea_month = {L: [0.0] * n for L in LINEAS}

    # Reparte la facturación real de contratos de cada cliente/mes entre sus
    # líneas de negocio, ponderando por el valor mensual de los contratos que
    # facturan ese mes (fallback: valor anual si ese mes no tiene flags activos).
    for p in panel_raw:
        real_arr = p.get("contr_meses_2026") or [0.0] * 12
        cons = contratos_by_cli.get(p["cliente"], [])
        for m in range(n):
            real_m = real_arr[m] if m < len(real_arr) else 0.0
            if abs(real_m) < 1:
                continue
            weights = defaultdict(float)
            for c in cons:
                if c["fact_flags"][m]:
                    weights[c["linea_negocio"]] += c["val_mes"]
            wsum = sum(weights.values())
            if wsum <= 0:
                weights = defaultdict(float)
                for c in cons:
                    weights[c["linea_negocio"]] += c["val"]
                wsum = sum(weights.values())
            if wsum <= 0:
                weights = {"Esterilización": 1.0}
                wsum = 1.0
            for L, w in weights.items():
                linea_month[L][m] += real_m * (w / wsum)

    # Escala cada mes para calzar exacto con contr_real_monthly (misma fuente que
    # "Ingresos por contratos" en EERR): corrige el pequeño drift de redondeo.
    for m in range(n):
        actual = sum(linea_month[L][m] for L in LINEAS)
        target = contr_real_monthly[m] if m < len(contr_real_monthly) else 0.0
        if abs(actual) > 1e-6:
            scale = target / actual
            for L in LINEAS:
                linea_month[L][m] *= scale

    # Otros ingresos = facturación real − ingreso por contratos, desglosado en
    # Trazabilidad + REAS (catálogos propios) y el resto de "Servicio Técnico"
    # (correctiva, repuestos sueltos, etc.) prorrateado por Marca.
    traz_mensual     = bbdd.get("traz_mensual", [0.0] * 12)
    reas_mensual     = bbdd.get("reas_mensual", [0.0] * 12)
    marca_mensual    = bbdd.get("marca_mensual", {})
    st_total_mensual = bbdd.get("st_total_mensual", [0.0] * 12)

    otros_total_month = [
        (fac_arr[m] if m < len(fac_arr) else 0.0) - (contr_real_monthly[m] if m < len(contr_real_monthly) else 0.0)
        for m in range(n)
    ]

    marca_month = {mk: [0.0] * n for mk in TOP_MARCAS_DESGLOSE + ["Otras Marcas"]}
    for m in range(n):
        tz = traz_mensual[m] if m < len(traz_mensual) else 0.0
        rs = reas_mensual[m] if m < len(reas_mensual) else 0.0
        remainder = otros_total_month[m] - tz - rs
        st_tot = st_total_mensual[m] if m < len(st_total_mensual) else 0.0
        if abs(st_tot) > 1e-6:
            used = 0.0
            for mk in TOP_MARCAS_DESGLOSE:
                serie = marca_mensual.get(mk, [0.0] * 12)
                amt = serie[m] if m < len(serie) else 0.0
                marca_month[mk][m] = remainder * (amt / st_tot)
                used += amt
            marca_month["Otras Marcas"][m] = remainder * ((st_tot - used) / st_tot)
        else:
            marca_month["Otras Marcas"][m] = remainder

    # Reclasificación de "Otros Ingresos" por tipo de reparación:
    # - Trazabilidad se fusiona con la marca ICTGroup (equipos de trazabilidad).
    # - Steelco (autoclaves) → 100% Reparación Esterilización.
    # - Pentax Medical (endoscopios) → 100% Reparación Endoscopía.
    # - El resto de marcas sin línea propia (TECSERVICE, Nacional, Otras Marcas)
    #   se prorratea usando las mismas proporciones mensuales de Ingreso por
    #   Contratos (Esterilización/Endoscopía/Dental); si un mes no tiene ingreso
    #   por contratos (ej. mes en curso sin datos aún), se usa la proporción
    #   promedio del período.
    total_linea_periodo = {L: sum(linea_month[L]) for L in LINEAS}
    total_linea_periodo_sum = sum(total_linea_periodo.values())
    fallback_prop = {
        L: (total_linea_periodo[L] / total_linea_periodo_sum if total_linea_periodo_sum > 0 else 1.0 / len(LINEAS))
        for L in LINEAS
    }

    remainder_pool_month = [
        marca_month["TECSERVICE"][m] + marca_month["NACIONAL"][m] + marca_month["Otras Marcas"][m]
        for m in range(n)
    ]
    reparacion_month = {L: [0.0] * n for L in LINEAS}
    for m in range(n):
        mes_total = sum(linea_month[L][m] for L in LINEAS)
        for L in LINEAS:
            prop = (linea_month[L][m] / mes_total) if mes_total > 1e-6 else fallback_prop[L]
            reparacion_month[L][m] = remainder_pool_month[m] * prop
    reparacion_month["Esterilización"] = [
        reparacion_month["Esterilización"][m] + marca_month["STEELCO"][m] for m in range(n)
    ]
    reparacion_month["Endoscopía"] = [
        reparacion_month["Endoscopía"][m] + marca_month["PENTAX MEDICAL"][m] for m in range(n)
    ]
    trazabilidad_month = [
        (traz_mensual[m] if m < len(traz_mensual) else 0.0) + marca_month["ICTGROUP"][m] for m in range(n)
    ]

    MM = 1_000_000.0
    def to_mm(arr):
        return [round(v / MM, 3) for v in arr]
    def with_total(arr):
        return arr + [round(sum(arr), 3)]

    meses_lbl = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio",
                 "Agosto","Septiembre","Octubre","Noviembre","Diciembre"][:n]

    contratos_lineas = [
        {"key": "esterilizacion", "label": "Esterilización", "valores": with_total(to_mm(linea_month["Esterilización"]))},
        {"key": "endoscopia",     "label": "Endoscopía",     "valores": with_total(to_mm(linea_month["Endoscopía"]))},
        {"key": "dental",         "label": "Dental",         "valores": with_total(to_mm(linea_month["Dental"]))},
    ]
    contratos_total = with_total(to_mm([contr_real_monthly[m] if m < len(contr_real_monthly) else 0.0 for m in range(n)]))

    otros_categorias = [
        {"key": "trazabilidad",              "label": "Trazabilidad",              "valores": with_total(to_mm(trazabilidad_month))},
        {"key": "reas",                       "label": "REAS",                      "valores": with_total(to_mm(reas_mensual[:n]))},
        {"key": "reparacion_esterilizacion", "label": "Reparación Esterilización", "valores": with_total(to_mm(reparacion_month["Esterilización"]))},
        {"key": "reparacion_endoscopia",      "label": "Reparación Endoscopía",     "valores": with_total(to_mm(reparacion_month["Endoscopía"]))},
        {"key": "reparacion_dental",          "label": "Reparación Dental",         "valores": with_total(to_mm(reparacion_month["Dental"]))},
    ]
    otros_total = with_total(to_mm(otros_total_month))

    total_general = with_total(to_mm([fac_arr[m] if m < len(fac_arr) else 0.0 for m in range(n)]))

    # ── Desglose por región ──────────────────────────────────────
    # Usa la MISMA fuente que la tabla global:
    #   - Contratos: contr_meses_2026 por cliente (= contr_real_monthly global)
    #   - Líneas: misma lógica de ponderación que linea_month global
    #   - Otros: prorrateo proporcional de otros_total_month global
    # Así los totales suman exacto a la tabla de desglose.
    # Región por cliente. Antes se cruzaba por nombre exacto, así que un
    # cliente escrito distinto entre FACTURACION y BASE MAPA caía en "Sin
    # región" — dejaba MM$38,6 de contratos ahí contra MM$21,6 facturados.
    # Ahora usa la misma normalización y búsqueda parcial que el lado de
    # facturación, para que ambos asignen igual.
    _reg_norm = {}
    for entry in (mapa_data or []):
        k = _norm_cli(entry.get("n", ""))
        if k:
            _reg_norm[k] = (entry.get("region") or "Sin región").strip() or "Sin región"

    def _region_de(nombre):
        k = _norm_cli(nombre)
        if k in _reg_norm:
            return _reg_norm[k]
        if len(k) >= 8:
            for mk, rg in _reg_norm.items():
                if len(mk) >= 8 and (mk in k or k in mk):
                    return rg
        return "Sin región"

    cli_to_region = {p["cliente"]: _region_de(p["cliente"]) for p in panel_raw}

    # Contratos y desglose por línea por región
    reg_con  = {}   # {region: [n floats]} contratos brutos
    reg_lin  = {}   # {region: {linea: [n floats]}} contratos por línea

    for p in panel_raw:
        cli    = p["cliente"]
        region = cli_to_region.get(cli, "Sin región")
        if region not in reg_con:
            reg_con[region] = [0.0] * n
            reg_lin[region] = {L: [0.0] * n for L in LINEAS}
        cons = contratos_by_cli.get(cli, [])
        ca   = p.get("contr_meses_2026") or [0.0] * 12
        for m in range(n):
            real_m = ca[m] if m < len(ca) else 0.0
            reg_con[region][m] += real_m
            if abs(real_m) < 1:
                continue
            # misma lógica de ponderación por línea que linea_month
            weights = defaultdict(float)
            for c in cons:
                if c["fact_flags"][m]:
                    weights[c["linea_negocio"]] += c["val_mes"]
            wsum = sum(weights.values())
            if wsum <= 0:
                weights = defaultdict(float)
                for c in cons:
                    weights[c["linea_negocio"]] += c["val"]
                wsum = sum(weights.values())
            if wsum <= 0:
                weights = {"Esterilización": 1.0}
                wsum = 1.0
            for L, w in weights.items():
                if L in reg_lin[region]:
                    reg_lin[region][L][m] += real_m * (w / wsum)

    # ── Acumulado Ene–mes_corte de 2025 y 2024 por región ───────────────────
    # Se replica exactamente el método del año en curso: los contratos van a la
    # región del cliente y "otros" se prorratea por el peso de contratos de cada
    # región. Así las tres columnas son comparables entre sí; mezclarlas con la
    # facturación real por cliente daría variaciones falsas.
    # Totales del año anterior con la misma base que usa el Resumen
    # (mensual_facturado = TS + Factura + los 3 catálogos de servicio).
    _mf = bbdd.get("mensual_facturado", {})
    def _tot_ano(y):
        return sum(v for m, v in _mf.get(y, {}).items() if m <= n)

    aa_cfg = []
    for y, campo in ((ANO - 1, "contr_2025"), (ANO - 2, "contr_2024")):
        con_reg = defaultdict(float)                       # contratos por región
        lin_reg = defaultdict(lambda: defaultdict(float))  # contratos por región y línea
        con_tot = 0.0
        for p in panel_raw:
            rg = cli_to_region.get(p["cliente"], "Sin región")
            v  = float(p.get(campo) or 0)
            if not v:
                continue
            con_reg[rg] += v
            con_tot     += v
            # Reparto por línea con el peso del valor anual de cada contrato,
            # que es el mismo criterio de respaldo que usa el año en curso
            # cuando no hay detalle mensual.
            pesos = defaultdict(float)
            for c in contratos_by_cli.get(p["cliente"], []):
                pesos[c["linea_negocio"]] += c["val"]
            tot_p = sum(pesos.values())
            if tot_p <= 0:
                pesos, tot_p = {"Esterilización": 1.0}, 1.0
            for L, w in pesos.items():
                lin_reg[rg][L] += v * (w / tot_p)
        # Facturación real por región de ese año, desde la BBDD por cliente:
        # mismo criterio que el año en curso, sin prorrateo.
        fac_reg = defaultdict(float)
        for cli_b, v2 in bbdd.get("ytd_cli_" + str(y), {}).items():
            fac_reg[_region_de(cli_b)] += float(v2 or 0)
        tot_ano = _tot_ano(y)
        aa_cfg.append((str(y), con_reg, lin_reg, con_tot, tot_ano, fac_reg))

    # ── "Otros Ingresos" a la región REAL de cada cliente ────────────────────
    # Antes se prorrateaba entre regiones según su peso en contratos, lo que
    # distorsionaba fuerte: Araucanía factura MM$242,6 casi todo fuera de
    # contrato y aparecía con MM$41,3, mientras Metropolitana se inflaba.
    # Ahora cada peso va donde se facturó, igual que hace el Mapa de Clientes.
    # Facturación mensual real por cliente (misma base que el resto del panel)
    _mpc = bbdd.get("mensual_por_cliente", {})
    reg_fac = defaultdict(lambda: [0.0] * n)
    for cli_b, arr in _mpc.items():
        rg = _region_de(cli_b)
        for m in range(n):
            reg_fac[rg][m] += arr[m] if m < len(arr) else 0.0
    for rg in reg_fac:
        reg_con.setdefault(rg, [0.0] * n)
        reg_lin.setdefault(rg, {L: [0.0] * n for L in LINEAS})

    reg_out = {}
    for r, con_arr in reg_con.items():
        con_mm = to_mm(con_arr)
        # Otros = facturación real de la región − sus contratos. Sin recortar
        # en cero: si los contratos devengados superan lo facturado el residuo
        # es negativo, y recortarlo inflaría el total de la región.
        fac_r   = reg_fac.get(r, [0.0] * n)
        otr_arr = [fac_r[m] - con_arr[m] for m in range(n)]
        otr_mm  = to_mm(otr_arr)
        tot_mm  = [round(con_mm[m] + otr_mm[m], 3) for m in range(n)]
        rlm     = reg_lin.get(r, {L: [0.0]*n for L in LINEAS})
        # Acumulados de años anteriores con el mismo prorrateo, abiertos por
        # línea, contratos, otros y total — igual que las filas del año actual.
        _aa = {}
        for ystr, con_reg, lin_reg, con_tot, tot_ano, fac_reg in aa_cfg:
            c_r = con_reg.get(r, 0.0)
            o_r = fac_reg.get(r, 0.0) - c_r
            _aa[ystr] = {
                "contratos": round(c_r / MM, 3),
                "otros":     round(o_r / MM, 3),
                "total":     round((c_r + o_r) / MM, 3),
                "lineas":    {L: round(lin_reg.get(r, {}).get(L, 0.0) / MM, 3) for L in LINEAS},
            }
        _a25 = _aa.get(str(ANO - 1), {})
        _a24 = _aa.get(str(ANO - 2), {})
        reg_out[r] = {
            "acum_2025":     _a25.get("total", 0.0),
            "acum_2024":     _a24.get("total", 0.0),
            "aa_2025":       _a25,
            "aa_2024":       _a24,
            "contratos": with_total(con_mm),
            "otros":     with_total(otr_mm),
            "total":     with_total(tot_mm),
            "lineas": {L: with_total(to_mm(rlm.get(L, [0.0]*n))) for L in LINEAS},
        }

    regiones_sorted = sorted(reg_out, key=lambda r: reg_out[r]["contratos"][n], reverse=True)

    return {
        "meses": meses_lbl,
        "contratos": {"lineas": contratos_lineas, "total": contratos_total},
        "otros":     {"categorias": otros_categorias, "total": otros_total},
        "total_general": total_general,
        "por_region": {"regiones": regiones_sorted, "data": reg_out},
    }


# ══════════════════════════════════════════════════════════════════════════════
# 7. ENSAMBLAR APP_DATA
# ══════════════════════════════════════════════════════════════════════════════
def build_app_data(contratos, panel_raw, bbdd, visitas, satisf, mes_corte, analisis_fac=None, base_instalada=None, mapa_data=None, gd_costo=None):

    # Lookups from CONTRATOS
    coord_by_cli   = {}
    presup_by_cli  = defaultdict(float)   # sum of val_anual for active COM contracts
    fin_contrato_by_cli  = {}  # última fecha de fin de contrato activo por cliente
    fin_fmt_by_cli       = {}
    inicio_fmt_by_cli    = {}
    dias_inicio_by_cli   = {}
    for c in contratos:
        if c["estado"] != "Activado":
            continue
        cli = c["cliente"]
        if c["coord"] and c["coord"] not in ("", "None"):
            coord_by_cli[cli] = c["coord"]
        if c["tipo"] == "Comercial":
            presup_by_cli[cli] += c["val"]
        # Guardar la fecha de fin más lejana (mayor cobertura para alertas)
        if cli not in fin_contrato_by_cli or c["fin"] > fin_contrato_by_cli[cli]:
            fin_contrato_by_cli[cli] = c["fin"]
            fin_fmt_by_cli[cli]      = c["fin_fmt"]
        # Guardar la fecha de inicio más antigua
        if cli not in inicio_fmt_by_cli or c["inicio"] < inicio_fmt_by_cli.get(cli + "_iso", "9999"):
            inicio_fmt_by_cli[cli]         = c["inicio_fmt"]
            inicio_fmt_by_cli[cli + "_iso"] = c["inicio"]
            dias_inicio_by_cli[cli]        = c["tpo_activo"]

    tipo_map    = bbdd["tipo_cli_map"]
    ytd_cli_25  = bbdd["ytd_cli_2025"]
    ytd_cli_24  = bbdd["ytd_cli_2024"]
    ytd_cli_26  = bbdd.get("ytd_cli_2026", {})

    # Clientes con contratos activos reales (para override de tiene_contrato)
    active_contract_clients = {c["cliente"] for c in contratos if c["estado"] == "Activado"}

    # Índice de facturación del año normalizado. El match exacto perdía
    # clientes por diferencias de tipografía entre FACTURACION y la BBDD
    # ("JOHNSON & JOHNSON" vs "Johnson y Johnson", "SPA" vs "S.p.A").
    _norm = _norm_cli
    _ytd26_idx = {}
    for _k, _v in bbdd.get("ytd_cli_2026", {}).items():
        _ytd26_idx[_norm(_k)] = _v

    # Build panel (enrich with CONTRATOS + BBDD data)
    panel = []
    for p in panel_raw:
        cli = p["cliente"]
        entry = {k: v for k, v in p.items() if not k.startswith("_")}
        entry["coord"]            = coord_by_cli.get(cli, "Sin contrato")
        entry["tipo_cli"]         = safe_str(tipo_map.get(cli, "")) or "Privado"
        entry["presup_contr_anio"] = round(presup_by_cli.get(cli, 0), 2)
        # Buscar en BBDD usando NombreAnalisis primero (coincide con col H del BBDD)
        nom = p.get("nombre_analisis") or cli
        entry["real_ytd_2025"]    = round(ytd_cli_25.get(nom, ytd_cli_25.get(cli, 0)))
        entry["real_ytd_2024"]    = round(ytd_cli_24.get(nom, ytd_cli_24.get(cli, 0)))
        # Facturación del año en curso con la misma base que "Ingresos
        # Totales" de Analisis Facturación: la suma de todos los clientes
        # cuadra con el KPI de portada.
        entry["real_ytd_fac"]     = round(_ytd26_idx.get(_norm(nom),
                                          _ytd26_idx.get(_norm(cli), 0)))
        entry["fin_contrato"]     = fin_contrato_by_cli.get(cli, "")
        entry["fin_fmt"]          = fin_fmt_by_cli.get(cli, "")
        entry["inicio_fmt"]       = inicio_fmt_by_cli.get(cli, "")
        entry["dias_inicio"]      = dias_inicio_by_cli.get(cli, None)
        # Override: tiene_contrato = True solo si hay contrato ACTIVO en CONTRATOS TODOS
        entry["tiene_contrato"]   = cli in active_contract_clients
        panel.append(entry)

    # ── Facturación del año por cliente, desde la BBDD ──────────────────────
    # El panel viene de la hoja FACTURACION y trae 111 clientes; la BBDD
    # factura además a clientes que no están ahí. Esta lista los incluye a
    # todos, así su suma cuadra con "Ingresos Totales" de Analisis Facturación.
    # cliente del panel (para tipo y contrato) indexado por nombre normalizado
    _pan_idx = {}
    for e in panel:
        _pan_idx[_norm(e["cliente"])] = e
        na = e.get("nombre_analisis")
        if na:
            _pan_idx.setdefault(_norm(na), e)
    _contr_norm = {_norm(c) for c in active_contract_clients}

    costo_cli = bbdd.get("costo_cli_2026", {})
    fact_clientes = []
    for cli_bbdd, monto in bbdd.get("ytd_cli_2026", {}).items():
        nom = safe_str(cli_bbdd).strip()
        if not nom or nom.lower() in ("nan", "none") or not monto:
            continue
        k  = _norm(nom)
        pe = _pan_idx.get(k)
        fact_clientes.append({
            "cliente":  pe["cliente"] if pe else nom,
            "tipo_cli": (pe.get("tipo_cli") if pe else None)
                        or safe_str(tipo_map.get(cli_bbdd, "")) or "Sin clasificar",
            "contrato": bool(pe.get("tiene_contrato")) if pe else (k in _contr_norm),
            "real":     round(float(monto)),
            "costo":    round(float(costo_cli.get(cli_bbdd, 0))),
        })
    # ── Costo de repuestos por cliente, desde la hoja GD ────────────────────
    # Reemplaza al costo de la BBDD y suma a los clientes que consumieron
    # repuestos sin facturar en el período: entran con real = 0 y margen
    # bruto negativo. La lista se arma en cada corrida, así que si mañana
    # esos clientes facturan salen solos de este grupo, y si aparecen otros
    # se incorporan sin tocar nada.
    if gd_costo:
        _vistos = set()
        for c in fact_clientes:
            k = _norm_cli(c["cliente"])
            _vistos.add(k)
            c["costo"] = gd_costo.get(k, {}).get("costo", 0)
            c["solo_costo"] = False
        # Los que consumieron repuestos sin facturar van consolidados en una
        # sola fila; el detalle queda en el tooltip.
        _sc = [e for k, e in gd_costo.items() if k not in _vistos and e["costo"]]
        _sc.sort(key=lambda e: -e["costo"])
        _nuevos = len(_sc)
        if _sc:
            fact_clientes.append({
                "cliente":  "Clientes sin Facturación",
                "tipo_cli": "Sin clasificar",
                "contrato": False,
                "real":     0,
                "costo":    sum(e["costo"] for e in _sc),
                "solo_costo": True,
                "n_sc":     _nuevos,
                "detalle":  [{"cliente": e["nombre"], "costo": e["costo"]} for e in _sc],
            })
        _kc = sum(x["costo"] for x in fact_clientes)
        _rc = sum(x["real"] for x in fact_clientes)
        print(f"       COSTO GD: {len(_vistos & set(gd_costo))} clientes con factura y costo | "
              f"{_nuevos} sólo costo | costo MM${_kc/1e6:,.1f} | "
              f"margen bruto MM${(_rc-_kc)/1e6:,.1f} ({(_rc-_kc)/_rc*100:.1f}%)")

    fact_clientes.sort(key=lambda x: -x["real"])
    _tot_fc = sum(x["real"] for x in fact_clientes)
    print(f"       FACT CLIENTES: {len(fact_clientes)} clientes | MM${_tot_fc/1e6:,.1f} "
          f"| costo MM${sum(x['costo'] for x in fact_clientes)/1e6:,.1f}")

    # ── Universo de la hoja Panel Facturación Cliente ───────────────────────
    # No sirve sumar `panel` directamente: la hoja FACTURACION trae filas
    # alias que apuntan al mismo cliente de la BBDD (CORPORACION DE
    # DESARROLLO SOCIAL DE BUIN → HOSPITAL SAN LUIS BUIN) y su facturación
    # se contaría dos veces, y además no cubre a todos los clientes que
    # factura la BBDD. Aquí se deduplica por el nombre con que la BBDD
    # identifica al cliente y se agregan los que faltan, para que el KPI de
    # la hoja dé los mismos MM$1.622,9 que la portada y que su propia tabla.
    pf_idx, panel_fact = {}, []
    for e in panel:
        k = _norm(e.get("nombre_analisis") or e["cliente"])
        prev = pf_idx.get(k)
        if prev is None:
            ent = dict(e)
            pf_idx[k] = ent
            pf_idx.setdefault(_norm(e["cliente"]), ent)
            panel_fact.append(ent)
            continue
        # Fila alias: los presupuestos se acumulan, la facturación es la
        # misma cifra de la BBDD y se deja una sola vez. Se conserva como
        # nombre visible el de la fila con más facturación propia.
        pf_idx.setdefault(_norm(e["cliente"]), prev)
        for f in ("presup_contr_ytd", "presup_contr_anio"):
            prev[f] = round((prev.get(f) or 0) + (e.get(f) or 0), 2)
        prev["tiene_contrato"] = bool(prev.get("tiene_contrato") or e.get("tiene_contrato"))
        if (e.get("real_ytd") or 0) > (prev.get("real_ytd") or 0):
            prev["cliente"] = e["cliente"]
            prev["tipo_cli"] = e.get("tipo_cli") or prev.get("tipo_cli")
        prev["real_ytd"] = max(prev.get("real_ytd") or 0, e.get("real_ytd") or 0)
        prev["real_ytd_fac"] = max(prev.get("real_ytd_fac") or 0, e.get("real_ytd_fac") or 0)

    # Clientes que factura la BBDD y que no tienen fila en FACTURACION
    for c in fact_clientes:
        if _norm(c["cliente"]) in pf_idx:
            continue
        ent = {
            "cliente":            c["cliente"],
            "nombre_analisis":    c["cliente"],
            "tipo_cli":           c["tipo_cli"],
            "tiene_contrato":     c["contrato"],
            "real_ytd_fac":       c["real"],
            "real_ytd":           c["real"],
            "real_ytd_2025":      0,
            "real_ytd_2024":      0,
            "presup_contr_ytd":   0.0,
            "presup_contr_anio":  0.0,
            "coord":              "Sin contrato",
            "fin_fmt":            "",
            "inicio_fmt":         "",
            "sin_fila_fact":      True,
        }
        pf_idx[_norm(c["cliente"])] = ent
        panel_fact.append(ent)

    _tot_pf = sum(e.get("real_ytd_fac", 0) for e in panel_fact)
    print(f"       PANEL FACT   : {len(panel_fact)} clientes | MM${_tot_pf/1e6:,.1f}"
          f" (panel crudo {len(panel)} = MM${sum(e.get('real_ytd_fac',0) for e in panel)/1e6:,.1f})")
    if abs(_tot_pf - _tot_fc) > 1:
        print(f"       ADVERTENCIA: panel_fact MM${_tot_pf/1e6:,.1f} != "
              f"fact_clientes MM${_tot_fc/1e6:,.1f}")

    # Monthly arrays helper
    def to_arr(d, year):
        yd = d.get(year, {})
        return [round(yd.get(m + 1, 0)) for m in range(12)]

    # presup_contr: expected monthly contract billing from CONTRATOS billing schedule
    presup_contr = [0.0] * 12
    for c in contratos:
        if c["estado"] == "Activado":
            for m in range(12):
                if c["fact_flags"][m]:
                    presup_contr[m] += c["val_mes"]
    presup_contr = [round(v, 2) for v in presup_contr]

    # YTD aggregates per year
    def ytd_agg(d, year):
        yd = d.get(year, {})
        tot_ytd  = round(sum(yd.get(m + 1, 0) for m in range(mes_corte)))
        tot_anio = round(sum(yd.get(m + 1, 0) for m in range(12)))
        return {"tot_ytd": tot_ytd, "tot_anio": tot_anio}

    mt = bbdd["mensual_total"]
    ytd = {
        str(ANO - 2): ytd_agg(mt, ANO - 2),
        str(ANO - 1): ytd_agg(mt, ANO - 1),
        str(ANO):     ytd_agg(mt, ANO),
    }

    # Contratos mensuales: suma de col AZ-BK (idx 51-62) de FACTURACIÓN por mes
    contr_real_monthly = [0.0] * 12
    for p in panel_raw:
        for m in range(12):
            contr_real_monthly[m] += p.get("contr_meses_2026", [0]*12)[m]
    contr_real_monthly = [round(v) for v in contr_real_monthly]
    # Otras = total facturado (BBDD filtrada) - contratos (mínimo 0)
    _fac_arr = to_arr(bbdd["mensual_facturado"], ANO)
    nocontr_real_monthly = [max(0, _fac_arr[m] - contr_real_monthly[m]) for m in range(12)]

    desglose_ingresos = compute_desglose_ingresos(contratos, panel_raw, bbdd, contr_real_monthly, _fac_arr, mes_corte, mapa_data=mapa_data)

    _MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
              "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

    _mes_nom = _MESES[mes_corte - 1] if 1 <= mes_corte <= 12 else "—"
    return {
        "mes_corte":        mes_corte,
        "mes_corte_nombre": _mes_nom,
        "periodo_nota":     f"Comparativo YTD: Ene–{_mes_nom} {ANO} vs Ene–{_mes_nom} {ANO-1} y {ANO-2}",
        "hoy":              TODAY.isoformat(),
        "ytd_contr_2024":   round(bbdd.get("ytd_contr_2024", 0)),
        "ytd_contr_2025":   round(bbdd.get("ytd_contr_2025", 0)),
        "panel":   panel,
        "panel_fact": panel_fact,
        "fact_clientes": fact_clientes,
        "mensual": {
            "total":        {str(y): to_arr(mt, y) for y in [ANO - 2, ANO - 1, ANO]},
            "facturado":    {str(y): to_arr(bbdd["mensual_facturado"], y) for y in [ANO - 2, ANO - 1, ANO]},
            "priv":         {str(y): to_arr(bbdd["mensual_priv"], y) for y in [ANO - 2, ANO - 1, ANO]},
            "pub":          {str(y): to_arr(bbdd["mensual_pub"],  y) for y in [ANO - 2, ANO - 1, ANO]},
            "presup_contr": presup_contr,
            "contr_real":   contr_real_monthly,
            "nocontr_real": nocontr_real_monthly,
            "por_ejecutivo": {
                eje: {str(y): to_arr(data, y) for y in [ANO - 2, ANO - 1, ANO]}
                for eje, data in bbdd.get("mensual_por_ejecutivo", {}).items()
            },
        },
        "ytd":                  ytd,
        "ppto_anual_total":     PPTO_ANUAL_TOTAL,
        "ppto_anual_contratos": round(PPTO_ANUAL_TOTAL * 0.5),
        "presup_mes_contr":     presup_contr,
        "satisf":               satisf,
        "visitas":              visitas,
        "analisis_fac":         analisis_fac or {},
        "base_instalada":       base_instalada or {"total":0,"por_linea":{},"por_tipo":[],"por_estado":{},"clientes":[]},
        "desglose_ingresos":    desglose_ingresos,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 7. CONSTRUIR DATA, NC_DATA, PERDIDOS_VG
# ══════════════════════════════════════════════════════════════════════════════
_DATA_EXCLUDE = {"fact_flags", "estado", "real_ytd", "real_2025", "real_2024"}

def _tiene_sucesor_reciente(contratos_cliente, ventana_dias=30):
    """True si, tras el contrato Expirado más reciente de un cliente, existe
    OTRO contrato (cualquier estado) cuyo inicio cae dentro de +/- ventana_dias
    respecto a esa fecha de término y que extiende la cobertura (termina
    después). Sirve para no marcar como "perdido" a un cliente que renovó
    pero cuyo contrato nuevo aún figura con Estado desactualizado en el Excel
    (mismo patrón de desfase que #198/#200 de Endoscopía)."""
    expirados = [c for c in contratos_cliente if c["estado"] != "Activado"]
    if not expirados:
        return False
    ultimo = max(expirados, key=lambda c: c["fin"])
    ultimo_fin = date.fromisoformat(ultimo["fin"])
    for c in contratos_cliente:
        if c is ultimo:
            continue
        c_fin = date.fromisoformat(c["fin"])
        if c_fin <= ultimo_fin:
            continue  # no extiende la cobertura del cliente
        c_inicio = date.fromisoformat(c["inicio"])
        if abs((c_inicio - ultimo_fin).days) <= ventana_dias:
            return True
    return False


def clientes_con_contrato_vigente(contratos):
    """Clientes con contrato REALMENTE vigente: Estado="Activado" en el Excel Y
    fecha de término aún no pasada.

    La columna Estado de CONTRATOS TODOS se actualiza a mano y queda atrasada:
    hay contratos marcados "Activado" cuya fecha venció hace semanas. Si sólo se
    mira Estado, un cliente que de verdad no continuó se cuenta como Renovado.
    Ver SUPUESTOS.txt punto 7a."""
    return {
        c["cliente"] for c in contratos
        if c["estado"] == "Activado" and c["dias_vence"] >= 0
    }


def corregir_estado_relacion(panel_raw, contratos):
    """Reconcilia el flag manual "No Continuó" de FACTURACIÓN con los contratos.

    Se aplica UNA sola vez sobre panel_raw, antes de construir APP_DATA.panel y
    DATA, para que todas las hojas muestren el mismo estado para un mismo
    cliente. Antes había dos fuentes divergentes: la tabla de detalle leía el
    flag crudo (Perdido) y Vencimientos/Visión General leían el corregido
    (Renovado), así que el mismo cliente aparecía distinto según la hoja."""
    vigentes = clientes_con_contrato_vigente(contratos)
    corregidos = []
    for p in panel_raw:
        if p.get("estado_relacion") == "Perdido" and p["cliente"] in vigentes:
            # El flag manual quedó atrasado: el cliente sí renovó y tiene un
            # contrato con fecha vigente.
            p["estado_relacion"] = "Renovado"
            p["_no_continuo"]    = False
            p["tiene_contrato"]  = True
            corregidos.append(p["cliente"])
    if corregidos:
        print(f"       Estado relacion corregido (No Continuo -> Renovado, tienen contrato vigente): {len(corregidos)}")
        for cli in corregidos:
            print(f"         - {cli}")
    return panel_raw


def build_data_arrays(contratos, panel_raw):
    # Lookup de estado_relacion por cliente desde FACTURACIÓN. Ya viene
    # reconciliado por corregir_estado_relacion(), así que se usa tal cual.
    panel_rel_map = {p["cliente"]: p.get("estado_relacion", "Nuevo") for p in panel_raw}
    contratos_by_cliente = defaultdict(list)
    for c in contratos:
        contratos_by_cliente[c["cliente"]].append(c)
    clientes_vigentes = clientes_con_contrato_vigente(contratos)

    # DATA: only active contracts
    data = []
    for c in contratos:
        if c["estado"] != "Activado":
            continue
        d = {k: v for k, v in c.items() if k not in _DATA_EXCLUDE}
        d["estado_relacion"] = panel_rel_map.get(c["cliente"], "Nuevo")
        data.append(d)

    # NC_DATA: new active contracts (started ≤90 days ago), all types
    nc_data = [d for d in data if d["es_nuevo"]]

    # PERDIDOS_VG: clients marked as no_continuo with no active contract
    perdidos = []
    for p in panel_raw:
        if not p.get("_no_continuo"):
            continue
        cli = p["cliente"]
        if cli in clientes_vigentes:
            continue  # Cliente renovó con otro contrato aún vigente
        if _tiene_sucesor_reciente(contratos_by_cliente.get(cli, [])):
            continue  # Renovó dentro de la ventana de 30 días; no está realmente perdido

        # Último contrato del cliente: el de término más reciente, sin importar
        # que el Excel lo siga marcando "Activado" (su fecha ya pasó).
        anteriores = [c for c in contratos if c["cliente"] == cli]
        last_c = max(anteriores, key=lambda c: c["fin"], default={}) if anteriores else {}

        perdidos.append({
            "n":              "—",
            "cliente":        cli,
            "coord":          last_c.get("coord", "—"),
            "tipo":           "Comercial",
            "inicio":         last_c.get("inicio", ""),
            "fin":            "",
            "inicio_fmt":     last_c.get("inicio_fmt", ""),
            "fin_fmt":        last_c.get("fin_fmt", ""),
            "val":            0,
            "val_mes":        0,
            "dias_vence":     -999,
            "long_dias":      0,
            "tpo_activo":     last_c.get("tpo_activo", 0),
            "pct_consumido":  100.0,
            "es_nuevo":       False,
            "estado_relacion": "Perdido",
            "fac_total":      p.get("real_anual_2024", 0) + p.get("real_anual_2025", 0) + p.get("real_ytd", 0),
            "fac_2026":       p.get("real_ytd", 0),
            "_es_perdido_fac": True,
            "fin_contrato":   last_c.get("fin", ""),
            "tipo_cli":       p.get("tipo_cli", "Privado"),
            "linea_negocio":  last_c.get("linea_negocio", "Esterilización"),
            "vendedor":       last_c.get("vendedor", "Sin vendedor"),
        })

    return data, nc_data, perdidos


# ══════════════════════════════════════════════════════════════════════════════
# 7b. BUILD ALERTA DATA — clientes con facturación bajo contrato
# ══════════════════════════════════════════════════════════════════════════════
def build_alerta_data(contratos, panel_raw, mes_corte):
    # Fuente: col AK = "SI", excluir garantías (todos los estados incluidos)
    by_cli = {}
    for c in contratos:
        if c.get("bajo_contrato") != "SI":
            continue
        if c["tipo"] == "Garantia":
            continue
        by_cli.setdefault(c["cliente"], []).append(c)

    # Mapa cliente → datos panel (para Real col H y Diferencia col D-H)
    panel_map = {p["cliente"]: p for p in panel_raw}

    result = []
    for cli, cli_contratos in by_cli.items():
        cli_contratos = sorted(cli_contratos, key=lambda x: x["inicio"])
        p = panel_map.get(cli, {})

        real_cli_total  = round(p.get("real_ytd", 0))          # col D FACTURACIÓN: total facturado al cliente
        esperado_cli    = round(p.get("presup_contr_ytd", 0))  # col H FACTURACIÓN: total contratado del cliente

        contracts_out = []
        for c in cli_contratos:
            contracts_out.append({
                "n":            c["n"],
                "coord":        c["coord"],
                "inicio_fmt":   c["inicio_fmt"],
                "fin_fmt":      c["fin_fmt"],
                "fact_flags":   c["fact_flags"],
                "n_mant_fecha": c.get("n_mant_actual", 0),   # col AF CONTRATOS TODOS
                "cuota_uf":     round(c.get("cuota_uf", 0), 2),
                "neta_mes":     round(c["val_mes"]),
                "n_mant":       c.get("n_mant", 0),
                "expected_ytd": round(c["real_ytd"]),         # col AG CONTRATOS TODOS
            })

        result.append({
            "cliente":        cli,
            "coord":          cli_contratos[0]["coord"],
            "total_expected": esperado_cli,                       # col H FACTURACIÓN
            "total_real":     real_cli_total,                     # col D FACTURACIÓN
            "total_gap":      esperado_cli - real_cli_total,      # col H − col D
            "contratos":      contracts_out,
        })

    result.sort(key=lambda x: -x["total_gap"])
    return result


# ══════════════════════════════════════════════════════════════════════════════
# 8. PARCHEAR BLOQUE DATA/APP_DATA EN EL HTML TEMPLATE
# ══════════════════════════════════════════════════════════════════════════════
def patch_html(html, data, app_data, mapa_data=None, casos_data=None, alerta_data=None):
    """Reemplaza el bloque <script>...const DATA=...;window.APP_DATA=...;</script>"""
    # Buscar "const DATA=" (con o sin espacio) y su bloque <script> contenedor
    data_idx = html.find("const DATA=")
    if data_idx == -1:
        data_idx = html.find("const DATA =")
    if data_idx == -1:
        print("  ADVERTENCIA: no se encontro 'const DATA' en el template.")
        return html

    # <script> más cercano ANTES de const DATA
    script_start = html.rfind("<script>", 0, data_idx)
    if script_start == -1:
        print("  ADVERTENCIA: no se encontro '<script>' antes de const DATA.")
        return html

    # Buscar window.APP_DATA después de const DATA (es la declaración usada en el HTML)
    app_idx = html.find("window.APP_DATA", data_idx)
    if app_idx == -1:
        print("  ADVERTENCIA: no se encontro 'window.APP_DATA' en el template.")
        return html

    # </script> inmediatamente después de window.APP_DATA
    script_end = html.find("</script>", app_idx)
    if script_end == -1:
        print("  ADVERTENCIA: no se encontro '</script>' despues de window.APP_DATA.")
        return html

    _MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
              "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    mes_nombre = _MESES[app_data["mes_corte"] - 1].upper()
    ano = app_data.get("hoy", "")[:4]

    mapa_json   = json.dumps(mapa_data   or [], ensure_ascii=False)
    casos_json  = json.dumps(casos_data  or {"casos": [], "equipos": []}, ensure_ascii=False)
    alerta_json = json.dumps(alerta_data or [], ensure_ascii=False)
    new_block = (
        "<script>\n"
        f"const DATA = {json.dumps(data, ensure_ascii=False)};\n\n"
        f"// ═══ DATOS ACTUALIZADOS A {mes_nombre} {ano} ═══\n"
        f"window.APP_DATA = {json.dumps(app_data, ensure_ascii=False)};\n"
        f"window.MAPA_DATA = {mapa_json};\n"
        f"window.CASOS_DATA = {casos_json};\n"
        f"window.ALERTA_DATA = {alerta_json};\n"
        "</script>"
    )
    return html[:script_start] + new_block + html[script_end + 9:]


# ══════════════════════════════════════════════════════════════════════════════
# 9. PARCHEAR datos.js (MES_CORTE, NC_DATA, PERDIDOS_VG)
# ══════════════════════════════════════════════════════════════════════════════
def patch_datos_js(js_content, mes_corte, nc_data, perdidos):
    # MES_CORTE
    js_content = re.sub(
        r'window\.MES_CORTE\s*=\s*\d+;',
        f"window.MES_CORTE = {mes_corte};",
        js_content,
    )
    # NC_DATA
    js_content = re.sub(
        r'const NC_DATA\s*=\s*\[.*?\];',
        f"const NC_DATA = {json.dumps(nc_data, ensure_ascii=False)};",
        js_content,
        flags=re.DOTALL,
    )
    # PERDIDOS_VG
    js_content = re.sub(
        r'const PERDIDOS_VG\s*=\s*\[.*?\];',
        f"const PERDIDOS_VG = {json.dumps(perdidos, ensure_ascii=False)};",
        js_content,
        flags=re.DOTALL,
    )
    return js_content


# ══════════════════════════════════════════════════════════════════════════════
# 10. INLINE TODOS LOS JS → HTML SELF-CONTAINED
# ══════════════════════════════════════════════════════════════════════════════
def build_final_html(html, datos_js_patched):
    for js_file in JS_FILES:
        # Build regex that matches <script src="file.js"></script>
        pattern = re.compile(
            rf'<script\s+src="{re.escape(js_file)}"\s*>\s*</script>',
            re.IGNORECASE,
        )
        if js_file == "datos.js":
            replacement = f"<script>\n{datos_js_patched}\n</script>"
        else:
            js_path = os.path.join(DIR, js_file)
            if not os.path.exists(js_path):
                print(f"  ADVERTENCIA: {js_file} no encontrado, se omite.")
                continue
            with open(js_path, encoding="utf-8") as f:
                js_content = f.read()
            replacement = f"<script>\n{js_content}\n</script>"

        # Usar lambda para evitar que re interprete backslashes del JS como escapes
        new_html = pattern.sub(lambda _: replacement, html)
        if new_html != html:
            print(f"  Inlined: {js_file}")
            html = new_html
        else:
            print(f"  ADVERTENCIA: no se encontro <script src=\"{js_file}\"> en el HTML.")

    return html


# ══════════════════════════════════════════════════════════════════════════════
# MAIN

# ══════════════════════════════════════════════════════════════════════════════
def _check_js_files():
    """Avisa si JS_FILES y $jsFiles de build.ps1 divergen.

    Son dos listas separadas: JS_FILES arma el standalone y $jsFiles arma
    index.html. Si un archivo está sólo en una, esa hoja funciona en un
    entregable y no existe en el otro, sin ningún error visible."""
    ps = os.path.join(DIR, "build.ps1")
    if not os.path.exists(ps):
        return
    with open(ps, encoding="utf-8", errors="ignore") as f:
        txt = f.read()
    m = re.search(r"\$jsFiles\s*=\s*@\((.*?)\)", txt, re.S)
    if not m:
        return
    en_ps  = set(re.findall(r"'([^']+\.js)'", m.group(1)))
    en_py  = set(JS_FILES)
    faltan_py = en_ps - en_py
    faltan_ps = en_py - en_ps
    if faltan_py:
        print(f"  AVISO: en build.ps1 pero NO en JS_FILES (faltarian en el standalone): {sorted(faltan_py)}")
    if faltan_ps:
        print(f"  AVISO: en JS_FILES pero NO en build.ps1 (faltarian en index.html): {sorted(faltan_ps)}")


def main():
    print("=" * 60)
    print("  EXTRACTOR DASHBOARD CONTRATOS TECSERVICE")
    print(f"  Excel : {os.path.basename(XLSX)}")
    print(f"  Fecha : {TODAY}  |  Ano : {ANO}")
    print("=" * 60)
    _check_js_files()

    # Detectar archivo: si existe .xlsb, convertir a xlsx temporal (siempre fresco)
    xlsx_to_use = XLSX
    if os.path.exists(XLSB):
        print(f"\n  Encontrado .xlsb — convirtiendo con Excel COM...")
        xlsx_to_use = _xlsb_to_xlsx(XLSB)
    elif not os.path.exists(XLSX):
        print(f"\nERROR: No se encontro el Excel en:\n  {XLSX}\n  ni en: {XLSB}")
        return

    if not os.path.exists(TMPL):
        print(f"\nERROR: No se encontro el template HTML en:\n  {TMPL}")
        return

    # ── Leer hojas simples con openpyxl ──────────────────────────────────────
    print("\n[1/5] Abriendo Excel con openpyxl...")
    wb = openpyxl.load_workbook(xlsx_to_use, read_only=True, data_only=True)

    print("[2/5] Leyendo CONTRATOS TODOS...")
    contratos = read_contratos(wb)
    activos   = [c for c in contratos if c["estado"] == "Activado"]
    expirados = [c for c in contratos if c["estado"] != "Activado"]
    print(f"       {len(activos)} activos | {len(expirados)} expirados/otros")

    print("[3/5] Leyendo FACTURACION...")
    panel_raw = read_facturacion(wb)
    print(f"       {len(panel_raw)} clientes en panel")

    print("[4/5] Leyendo VISITAS, SATISFACCION y BASE MAPA...")
    # Necesitamos mes_corte antes de procesar visitas; hacemos un pase rápido
    # leyendo BBDD primero (lo haremos abajo) — usamos mes=MES_CORTE provisional
    satisf = read_satisfaccion(wb)
    wb.close()
    print(f"       Satisfaccion: {satisf['global']['n']} respuestas | NPS {satisf['nps']['nps']:+}")

    # ── Leer BBDD FACTURACION con pandas ─────────────────────────────────────
    print("[5/5] Leyendo BBDD FACTURACION...")
    bbdd      = read_bbdd(xlsx_to_use)
    mes_corte = bbdd["mes_corte"]
    print(f"       MES_CORTE detectado automaticamente: {mes_corte}")

    # Ahora sí leemos visitas con mes_corte correcto
    wb2 = openpyxl.load_workbook(xlsx_to_use, read_only=True, data_only=True)
    visitas = read_visitas(wb2, mes_corte)
    analisis_fac = read_analisis_fac(wb2)
    base_instalada = read_base_instalada(wb2)
    mapa_data  = read_mapa(wb2)
    casos_data = read_casos(wb2)
    ratios2 = read_ratios2(wb2)
    resumen_programas = read_resumen_tipos_programas(wb2)
    inv_ts = read_inventario_ts(wb2)
    rep_vend = read_repuestos_vendidos(wb2)
    eq_fallas = read_equipos_fallas(wb2)
    prosp_bi  = read_prospectos_bi(wb2)
    fact_desg = read_fact_desglose(wb2)
    br_oport = read_brecha_oport(wb2)
    br_stock = read_brecha_stock(wb2)
    cli_rel  = read_clientes_relevantes(wb2)
    back_ord = read_back_order_idx(wb2)
    pipe_st  = read_pipeline_st(wb2)
    gd_costo = read_gd_costos(wb2)
    wb2.close()
    eg = visitas["resumen"].get("Eglys Ramirez", {})
    cr = visitas["resumen"].get("Cristian Perez", {})
    print(f"       Visitas YTD: Eglys {eg.get('tot_2026_ytd',0)} | Cristian {cr.get('tot_2026_ytd',0)}")
    ts_ytd = analisis_fac.get("ts_total_ytd", 0)
    ts_ing = analisis_fac.get("ts_ingresos", 0)
    print(f"       Analisis Fac: Ingresos TS MM${ts_ing/1e6:.1f} | Total YTD MM${ts_ytd/1e6:.1f}")
    _sin_geo = sum(1 for c in mapa_data if c.get("lat") is None or c.get("lon") is None)
    print(f"       BASE MAPA: {len(mapa_data)} clientes"
          + (f" ({_sin_geo} sin coordenadas, no se dibujan en el mapa)" if _sin_geo else ""))

    # ── Construir estructuras de datos ───────────────────────────────────────
    print("[6/6] Construyendo estructuras de datos...")
    # Reconciliar el flag manual "No Continuó" con los contratos ANTES de
    # derivar APP_DATA.panel y DATA, para que ambas fuentes coincidan.
    panel_raw = corregir_estado_relacion(panel_raw, contratos)
    app_data = build_app_data(contratos, panel_raw, bbdd, visitas, satisf, mes_corte, analisis_fac, base_instalada, mapa_data=mapa_data, gd_costo=gd_costo)
    app_data["ratios2"] = ratios2
    app_data["resumen_programas"] = resumen_programas
    app_data["inv_ts"] = inv_ts
    app_data["rep_vend"] = rep_vend
    app_data["eq_fallas"] = eq_fallas
    app_data["prosp_bi"] = prosp_bi
    app_data["fact_desglose"] = fact_desg
    app_data["br_oport"] = br_oport
    app_data["br_stock"] = br_stock
    app_data["cli_rel"]  = cli_rel
    app_data["back_order"] = back_ord
    app_data["pipeline_st"] = pipe_st
    # Costo GD que no queda asignado a ningún cliente de la tabla
    if gd_costo:
        _hu = {_norm_cli(c["cliente"]) for c in app_data.get("fact_clientes", [])}
        app_data["gd_costo_fuera"] = round(sum(e["costo"] for k, e in gd_costo.items() if k not in _hu))
        app_data["gd_costo_total"] = round(sum(e["costo"] for e in gd_costo.values()))
    if pipe_st:
        print(f"       PIPELINE ST: {pipe_st['n']} oportunidades | {pipe_st['n_clientes']} clientes | "
              f"{len(pipe_st['lineas'])} lineas | anios {','.join(pipe_st['anios'])} | "
              f"MM${pipe_st['monto_tot']/1e6:,.1f}")
    if cli_rel:
        print(f"       CLIENTES RELEVANTES: {cli_rel['n_clientes']} clientes | "
              f"6m ({cli_rel['periodo6']}) MM${cli_rel['tot_m6']/1e6:,.1f} | "
              f"12m ({cli_rel['periodo12']}) MM${cli_rel['tot_m12']/1e6:,.1f} | "
              f"back order {len(back_ord)} SKU")
    if inv_ts:
        print(f"       INVENTARIO TS: {inv_ts['n_marcas']} marcas | {inv_ts['total_skus']} SKUs | "
              f"{inv_ts['total_stock']:,.0f} un | MM${inv_ts['total_costo']/1e6:,.1f}")
    if rep_vend:
        print(f"       REPUESTOS VENDIDOS: {len(rep_vend['marcas'])} marcas | {len(rep_vend['meses'])} meses "
              f"({rep_vend['meses'][0]['lbl']}-{rep_vend['meses'][-1]['lbl']}) | "
              f"{rep_vend['n_clientes']} clientes | {len(rep_vend['familias'])} familias | MM${rep_vend['tot_monto_g']/1e6:,.1f} | "
              f"{rep_vend['tot_cant_g']:,.0f} un")
    # Momento real en que se generó el panel, en hora de Chile. Antes se
    # forzaba la hora del proceso batch (02:50 am) y, si la corrida era más
    # tarde, se avanzaba un día: el tablero terminaba mostrando una fecha en el
    # futuro. La zona horaria se fija explícitamente para que el sello sea el
    # mismo aunque el panel se genere desde un equipo en otro huso.
    try:
        from zoneinfo import ZoneInfo
        _ahora = datetime.now(ZoneInfo("America/Santiago"))
    except Exception:
        _ahora = datetime.now()
    app_data["actualizado_label"] = formato_actualizacion(_ahora)
    app_data["actualizado_iso"] = _ahora.isoformat()
    print(f"       {app_data['actualizado_label']}")
    data, nc_data, perdidos = build_data_arrays(contratos, panel_raw)
    total_com_val = sum(d["val"] for d in data if d["tipo"] == "Comercial")
    total_gar_val = sum(d["val"] for d in data if d["tipo"] == "Garantia")
    print(f"       DATA: {len(data)} contratos | Cartera COM: MM${total_com_val/1e6:.1f} | GAR: MM${total_gar_val/1e6:.1f} | Total: MM${(total_com_val+total_gar_val)/1e6:.1f}")
    print(f"       NC_DATA: {len(nc_data)} nuevos | PERDIDOS: {len(perdidos)}")
    alerta_data = build_alerta_data(contratos, panel_raw, mes_corte)
    print(f"       ALERTA_DATA: {len(alerta_data)} clientes con facturación bajo contrato")

    # Brecha por facturación bajo contrato: no viene de una hoja propia, se
    # deriva de ALERTA_DATA (facturación esperada por contrato − facturación
    # real), así queda siempre alineada con la hoja "Bajo Contrato" del panel.
    app_data["br_contrato"] = {
        "total":     round(sum(c.get("total_gap", 0) for c in alerta_data)),
        "esperado":  round(sum(c.get("total_expected", 0) for c in alerta_data)),
        "real":      round(sum(c.get("total_real", 0) for c in alerta_data)),
        "n_clientes": len(alerta_data),
    }
    _b = app_data["br_contrato"]
    print(f"       BRECHAS: oport MM${(br_oport.get('total',0))/1e6:,.1f} | "
          f"contrato MM${_b['total']/1e6:,.1f} | stock MM${(br_stock.get('total',0))/1e6:,.1f}")
    enrich_mapa_data(mapa_data, contratos, satisf)
    completar_mapa(mapa_data, app_data.get("fact_clientes") or [])
    cc_count = sum(1 for c in mapa_data if c["cc"])
    print(f"       MAPA_DATA: {len(mapa_data)} clientes | {cc_count} con contrato")
    print(f"       CASOS: {len(casos_data['casos'])} casos relevantes | {len(casos_data['equipos'])} equipos detenidos")

    # ── Parchear template.html (fuente de build.ps1) ─────────────────────────
    print("\nParcheando template.html...")
    with open(TMPL, encoding="utf-8") as f:
        html = f.read()
    html = patch_html(html, data, app_data, mapa_data, casos_data, alerta_data)
    with open(TMPL, "w", encoding="utf-8") as f:
        f.write(html)
    print("  Actualizado: template.html")

    # ── Parchear y guardar datos.js ───────────────────────────────────────────
    datos_js_path = os.path.join(DIR, "datos.js")
    with open(datos_js_path, encoding="utf-8") as f:
        datos_js = f.read()
    datos_js_patched = patch_datos_js(datos_js, mes_corte, nc_data, perdidos)
    with open(datos_js_path, "w", encoding="utf-8") as f:
        f.write(datos_js_patched)
    print("  Actualizado: datos.js")

    # ── Generar dashboard standalone (v16.2) ──────────────────────────────────
    print("\nGenerando dashboard standalone...")
    html_final = build_final_html(html, datos_js_patched)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(html_final)

    size_kb = round(os.path.getsize(OUT) / 1024)
    print()
    print("=" * 60)
    print("  LISTO")
    print(f"  template.html  : actualizado con datos nuevos")
    print(f"  datos.js       : actualizado (MES_CORTE, NC_DATA, PERDIDOS)")
    print(f"  Standalone     : {os.path.basename(OUT)} ({size_kb} KB)")
    print(f"  MES_CORTE : {mes_corte}")
    print(f"  Contratos activos : {len(data)}")
    print(f"  Clientes panel    : {len(app_data['panel'])}")
    print(f"  Cartera COM       : MM${total_com_val/1e6:.1f}")
    print(f"  Cartera GAR       : MM${total_gar_val/1e6:.1f}")
    print(f"  Cartera Total     : MM${(total_com_val+total_gar_val)/1e6:.1f}")
    print()
    print("  SIGUIENTE PASO:")
    print(r"  cd src && .\build.ps1")
    print("=" * 60)


if __name__ == "__main__":
    main()
