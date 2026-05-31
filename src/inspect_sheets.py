import openpyxl

XLSX = '../data/CONTRATOS -FACTURACION -SATISFACCION -VISITAS.xlsx'
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
print("Hojas:", wb.sheetnames)

# Buscar hoja de analisis (con acento)
analisis_name = None
gdppto_name = None
for name in wb.sheetnames:
    nl = name.lower().replace('ó','o').replace('ó','o')
    if 'analisis' in nl or 'análisis' in name.lower():
        analisis_name = name
    if 'gd' in nl and 'ppto' in nl:
        gdppto_name = name

print("Analisis sheet:", analisis_name)
print("GD-PPTO sheet:", gdppto_name)
print()

if analisis_name:
    print("=== " + analisis_name + " (filas 1-40, cols A-P) ===")
    ws = wb[analisis_name]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), 1):
        vals = list(row[:16])
        if any(v is not None for v in vals):
            clean = [round(v,0) if isinstance(v,float) else v for v in vals]
            print("F" + str(i).zfill(2) + ": " + str(clean))

print()
if gdppto_name:
    print("=== " + gdppto_name + " (filas 1-30, cols A-N) ===")
    ws2 = wb[gdppto_name]
    for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        vals = list(row[:14])
        if any(v is not None for v in vals):
            clean = [round(v,0) if isinstance(v,float) else v for v in vals]
            print("F" + str(i).zfill(2) + ": " + str(clean))

wb.close()
