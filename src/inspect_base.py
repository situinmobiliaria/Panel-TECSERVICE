import openpyxl

XLSX = '../data/CONTRATOS -FACTURACION -SATISFACCION -VISITAS.xlsx'
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

ws = wb['BASE INSTALADA']
print("=== BASE INSTALADA — Encabezados (fila 1) ===")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=1, values_only=True), 1):
    for idx, col in enumerate(row):
        if col is not None:
            print(f"  col[{idx}]: {col}")

print()
print("=== BASE INSTALADA — Primeras 8 filas de datos ===")
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=9, values_only=True), 2):
    vals = [round(v,1) if isinstance(v,float) else v for v in row]
    print(f"F{i:02d}: {vals}")

print()
print(f"Total filas estimadas: {ws.max_row}")
wb.close()
